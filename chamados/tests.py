from io import BytesIO
from tempfile import gettempdir
from unittest.mock import AsyncMock, patch

from asgiref.sync import async_to_sync
from channels.routing import URLRouter
from channels.db import database_sync_to_async
from channels.testing import WebsocketCommunicator
from django.contrib.auth.models import Group, User
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files.uploadedfile import (
    InMemoryUploadedFile,
    SimpleUploadedFile,
    TemporaryUploadedFile,
)
from django.core.handlers.asgi import ASGIHandler
from django.test import SimpleTestCase, TestCase, TransactionTestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from chamados.lider_service import InventarioLiderService
from chamados.models import (
    AliasUsuario,
    CategoriaChamado,
    Chamado,
    ChamadoAnexo,
    ChamadoAvaliacao,
    ChamadoConexaoAtendente,
    ChamadoEvento,
    ChamadoMensagem,
    ChamadoSessaoAtendimento,
    PendenciaVinculoLider,
    validar_assinatura_anexo,
    validar_mime_anexo,
    validar_tamanho_anexo,
)
from chamados.policies import ChamadoAccessPolicy, GruposChamados
from chamados.routing import websocket_urlpatterns
from chamados.services import ChamadoService
from estoque.models import Base, Comunicado, Empresa, Equipamento, Perfil, Produto, Sick
from insumos.models import Cliente, Inventario
from ordens_servico.models import OrdemServico


class ChamadosUploadASGITests(SimpleTestCase):
    @override_settings(FILE_UPLOAD_MAX_MEMORY_SIZE=1024 * 1024)
    def test_corpo_asgi_acima_do_limiar_e_movido_para_disco(self):
        mensagens = iter([
            {
                'type': 'http.request',
                'body': b'x' * (1024 * 1024 + 1),
                'more_body': False,
            },
        ])

        async def receber():
            return next(mensagens)

        arquivo = async_to_sync(ASGIHandler().read_body)(receber)
        try:
            self.assertTrue(arquivo._rolled)
        finally:
            arquivo.close()


class ChamadosIntegracaoTests(TestCase):
    def setUp(self):
        self.empresa = Empresa.objects.create(nome='Empresa Principal')
        self.base = Base.objects.create(empresa=self.empresa, nome='Base Curitiba')
        self.outra_base = Base.objects.create(empresa=self.empresa, nome='Base Recife')
        self.solicitante = User.objects.create_user('solicitante', password='SenhaForte123!')
        self.solicitante.perfil.empresa = self.empresa
        self.solicitante.perfil.role = Perfil.Role.OPERADOR
        self.solicitante.perfil.save()
        self.solicitante.perfil.regionais.add(self.base)
        self.cliente = Cliente.objects.create(sigla='CLI', nome='Cliente Teste')
        self.inventario = Inventario.objects.create(
            cliente=self.cliente,
            loja='Loja Centro',
            base=self.base,
            data_inicio=timezone.localdate(),
            criado_por=self.solicitante,
            lider='Maria Souza',
            lider_usuario=self.solicitante,
        )
        self.outro = User.objects.create_user('outro', password='SenhaForte123!')
        self.outro.perfil.empresa = self.empresa
        self.outro.perfil.role = Perfil.Role.OPERADOR
        self.outro.perfil.save()
        self.outro.perfil.regionais.add(self.outra_base)
        self.admin = User.objects.create_user('admin_chamados', password='SenhaForte123!')
        self.admin.perfil.empresa = self.empresa
        self.admin.perfil.role = Perfil.Role.ADMIN
        self.admin.perfil.save()
        self.atendente = User.objects.create_user('atendente', password='SenhaForte123!')
        self.atendente.perfil.empresa = self.empresa
        self.atendente.perfil.regionais.add(self.base)
        self.atendente.perfil.save()
        grupo_suporte, _ = Group.objects.get_or_create(name=GruposChamados.SUPORTE)
        self.atendente.groups.add(grupo_suporte)
        self.supervisor = User.objects.create_user('supervisor', password='SenhaForte123!')
        self.supervisor.perfil.empresa = self.empresa
        self.supervisor.perfil.regionais.add(self.base)
        self.supervisor.perfil.save()
        grupo_supervisor, _ = Group.objects.get_or_create(name=GruposChamados.SUPERVISOR)
        self.supervisor.groups.add(grupo_supervisor)
        self.categoria, _ = CategoriaChamado.objects.get_or_create(
            nome='ROUTER NÃO FUNCIONA', defaults={'sla_horas': 4},
        )
        self.produto = Produto.objects.create(
            codigo='RTR-1', descricao='Router', fabricante='Cisco', modelo='R1', categoria='Routers'
        )
        self.equipamento = Equipamento.objects.create(
            produto=self.produto, numero_serie='SER-CH-1', patrimonio='PAT-CH-1',
            codigo='EQ-CH-1', regional=self.base,
        )

    def abrir(self):
        return ChamadoService.abrir(
            usuario=self.solicitante,
            base=self.base,
            inventario=self.inventario,
            equipamento=self.equipamento,
            categoria=self.categoria,
            loja='Loja Centro',
            lider='Maria Souza',
            titulo='Sem conexão no inventário',
            descricao='Router não liga e a equipe está parada.',
            prioridade=Chamado.Prioridade.CRITICA,
        )

    def resolver_para_avaliacao(self, chamado=None):
        chamado = chamado or self.abrir()
        ChamadoService.assumir(chamado, self.atendente)
        ChamadoService.resolver(
            chamado,
            self.atendente,
            causa_raiz='Travamento de firmware.',
            solucao='Equipamento reiniciado e conexão restabelecida.',
        )
        chamado.refresh_from_db()
        return chamado

    def test_abertura_usa_usuarios_bases_e_comunicados_do_gerenciador(self):
        chamado = self.abrir()

        self.assertRegex(chamado.protocolo, r'^CH-\d{4}-\d{4}-000001$')
        self.assertEqual(chamado.titulo, 'SEM CONEXÃO NO INVENTÁRIO')
        self.assertEqual(chamado.loja, 'LOJA CENTRO')
        self.assertEqual(chamado.empresa, self.empresa)
        self.assertEqual(chamado.base, self.base)
        comunicado = Comunicado.objects.get(dados__chamado_id=chamado.pk)
        self.assertTrue(comunicado.usuarios.filter(pk=self.admin.pk).exists())
        self.assertTrue(comunicado.usuarios.filter(pk=self.solicitante.pk).exists())

    def test_chamado_registra_se_ocorreu_antes_ou_durante_o_inventario(self):
        antes = self.abrir()
        self.assertEqual(antes.momento_inventario_abertura, Chamado.MomentoInventario.ANTES)
        self.inventario.status = 'EM_ANDAMENTO'
        self.inventario.save(update_fields=['status'])
        durante = self.abrir()
        self.assertEqual(
            durante.momento_inventario_abertura,
            Chamado.MomentoInventario.EM_ANDAMENTO,
        )

    def test_abertura_exibe_somente_atendentes_online(self):
        ChamadoConexaoAtendente.objects.create(
            usuario=self.atendente, canal='teste-lista-atendentes-online'
        )
        self.client.force_login(self.solicitante)
        resposta = self.client.get(reverse('chamados:criar'))
        self.assertEqual(resposta.status_code, 200)
        self.assertContains(resposta, self.atendente.username)
        self.assertNotContains(resposta, self.supervisor.username)

    def test_isola_chamado_de_usuario_sem_acesso(self):
        chamado = self.abrir()
        self.client.force_login(self.outro)

        response = self.client.get(reverse('chamados:detalhe', args=[chamado.pk]))

        self.assertEqual(response.status_code, 404)

    def test_detalhe_exibe_documentacao_vinculada_ao_produto_sem_inferir_solucao(self):
        self.produto.codigo = 'IMP-BR-XR'
        self.produto.fabricante = 'Xerox'
        self.produto.modelo = 'Phaser 3020'
        self.produto.categoria = 'Impressoras'
        self.produto.save()
        chamado = self.abrir()
        self.client.force_login(self.solicitante)

        response = self.client.get(reverse('chamados:detalhe', args=[chamado.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Documentação do equipamento')
        self.assertContains(response, 'Xerox Phaser 3020 - Guia interno')
        self.assertContains(response, 'Guia do usuário Phaser 3020')
        self.assertContains(response, 'Confirme o diagnóstico')
        self.assertContains(response, 'rel="noopener noreferrer"')

    def test_atendente_assume_conversa_e_resolve_com_historico(self):
        chamado = self.abrir()
        ChamadoService.assumir(chamado, self.atendente)
        mensagem = ChamadoService.adicionar_mensagem(
            chamado, self.atendente, 'Reiniciamos o equipamento e validamos a rede.'
        )
        ChamadoService.resolver(
            chamado, self.atendente,
            causa_raiz='Travamento de firmware.',
            solucao='Equipamento reiniciado e conexão restabelecida.',
        )
        chamado.refresh_from_db()

        self.assertEqual(chamado.atendente, self.atendente)
        self.assertEqual(chamado.status, Chamado.Status.AVALIACAO)
        self.assertEqual(mensagem.texto, 'REINICIAMOS O EQUIPAMENTO E VALIDAMOS A REDE.')
        self.assertGreaterEqual(chamado.eventos.count(), 4)

    def test_evento_websocket_identifica_solicitante_para_destaque_da_avaliacao(self):
        chamado = self.abrir()
        camada = type('CamadaTeste', (), {})()
        camada.group_send = AsyncMock()

        with (
            patch('chamados.services.get_channel_layer', return_value=camada),
            self.captureOnCommitCallbacks(execute=True),
        ):
            ChamadoService._evento(
                chamado,
                'RESOLUCAO',
                'CHAMADO RESOLVIDO E ENVIADO PARA AVALIAÇÃO.',
                self.atendente,
            )

        payloads = [chamada.args[1]['payload'] for chamada in camada.group_send.await_args_list]
        self.assertTrue(payloads)
        self.assertTrue(all(item['solicitante_id'] == self.solicitante.pk for item in payloads))

    def test_atendente_assume_pelo_formulario_sem_interceptacao_da_avaliacao(self):
        chamado = self.abrir()
        self.client.force_login(self.atendente)
        detalhe_url = reverse('chamados:detalhe', args=[chamado.pk])
        assumir_url = reverse('chamados:assumir', args=[chamado.pk])

        pagina = self.client.get(detalhe_url)

        self.assertEqual(pagina.status_code, 200)
        self.assertContains(pagina, 'id="assumir-form"')
        self.assertContains(pagina, 'id="ativar-alertas-chamados"')
        self.assertContains(pagina, f'action="{assumir_url}"')
        self.assertNotContains(pagina, 'id="avaliacao-form"')

        resposta = self.client.post(assumir_url)

        self.assertRedirects(resposta, detalhe_url)
        chamado.refresh_from_db()
        self.assertEqual(chamado.atendente, self.atendente)
        self.assertEqual(chamado.status, Chamado.Status.EM_ATENDIMENTO)

    def test_solicitante_ve_modal_e_aviso_persistente_quando_pode_avaliar(self):
        chamado = self.resolver_para_avaliacao()
        self.client.force_login(self.solicitante)

        pagina = self.client.get(reverse('chamados:detalhe', args=[chamado.pk]))

        self.assertTrue(pagina.context['pode_avaliar'])
        self.assertContains(pagina, 'id="avaliacaoModal"', count=1)
        self.assertContains(pagina, 'id="avaliacao-form"', count=1)
        self.assertContains(pagina, 'id="avaliacao-feedback"', count=1)
        self.assertContains(pagina, 'Este atendimento ainda aguarda sua avaliação.')
        self.assertContains(pagina, chamado.resolucao)
        self.assertContains(pagina, 'role="radiogroup"')
        self.assertContains(pagina, 'data-auto-show="true"')

    def test_atendente_e_terceiro_nao_podem_avaliar(self):
        chamado = self.resolver_para_avaliacao()
        url = reverse('chamados:avaliar', args=[chamado.pk])
        headers = {'HTTP_X_REQUESTED_WITH': 'XMLHttpRequest'}

        for usuario in (self.atendente, self.admin):
            with self.subTest(usuario=usuario.username):
                self.client.force_login(usuario)
                pagina = self.client.get(reverse('chamados:detalhe', args=[chamado.pk]))
                self.assertFalse(pagina.context['pode_avaliar'])
                self.assertNotContains(pagina, 'id="avaliacao-form"')
                resposta = self.client.post(
                    url,
                    {'nota': 5, 'resolvido': 'True', 'comentario': ''},
                    **headers,
                )
                self.assertEqual(resposta.status_code, 400)
                self.assertIn('SOMENTE O SOLICITANTE', resposta.json()['erro'])
        self.assertFalse(ChamadoAvaliacao.objects.filter(chamado=chamado).exists())

    def test_endpoint_rejeita_notas_invalidas_e_avaliacao_duplicada(self):
        chamado = self.resolver_para_avaliacao()
        url = reverse('chamados:avaliar', args=[chamado.pk])
        headers = {'HTTP_X_REQUESTED_WITH': 'XMLHttpRequest'}
        self.client.force_login(self.solicitante)

        for nota in (0, 6):
            with self.subTest(nota=nota):
                resposta = self.client.post(
                    url,
                    {'nota': nota, 'resolvido': 'True', 'comentario': ''},
                    **headers,
                )
                self.assertEqual(resposta.status_code, 400)
                self.assertIn('nota', resposta.json()['erros'])

        resposta = self.client.post(
            url,
            {'nota': 5, 'resolvido': 'True', 'comentario': 'Ótimo atendimento.'},
            **headers,
        )
        self.assertEqual(resposta.status_code, 200)
        self.assertTrue(resposta.json()['encerrado'])
        chamado.refresh_from_db()
        self.assertEqual(chamado.status, Chamado.Status.ENCERRADO)
        self.assertEqual(ChamadoAvaliacao.objects.filter(chamado=chamado).count(), 1)

        resposta_duplicada = self.client.post(
            url,
            {'nota': 4, 'resolvido': 'True', 'comentario': ''},
            **headers,
        )
        self.assertEqual(resposta_duplicada.status_code, 400)
        self.assertIn('NÃO ESTÁ AGUARDANDO AVALIAÇÃO', resposta_duplicada.json()['erro'])
        self.assertEqual(ChamadoAvaliacao.objects.filter(chamado=chamado).count(), 1)

    def test_endpoint_avaliacao_negativa_reabre_chamado(self):
        chamado = self.resolver_para_avaliacao()
        self.client.force_login(self.solicitante)

        resposta = self.client.post(
            reverse('chamados:avaliar', args=[chamado.pk]),
            {'nota': 2, 'resolvido': 'False', 'comentario': 'Ainda sem conexão.'},
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        self.assertEqual(resposta.status_code, 200)
        self.assertFalse(resposta.json()['encerrado'])
        chamado.refresh_from_db()
        self.assertEqual(chamado.status, Chamado.Status.REABERTO)

    def test_nota_interna_nao_e_exibida_nem_gera_comunicado(self):
        chamado = self.abrir()
        ChamadoService.assumir(chamado, self.atendente)
        comunicados_antes = Comunicado.objects.filter(dados__chamado_id=chamado.pk).count()
        ChamadoService.adicionar_mensagem(
            chamado, self.atendente, 'Possível troca em garantia.', nota_interna=True
        )
        self.client.force_login(self.solicitante)

        response = self.client.get(reverse('chamados:detalhe', args=[chamado.pk]))

        self.assertNotContains(response, 'POSSÍVEL TROCA EM GARANTIA')
        self.assertFalse(response.context['mensagens_chamado'].filter(nota_interna=True).exists())
        self.assertEqual(
            Comunicado.objects.filter(dados__chamado_id=chamado.pk).count(),
            comunicados_antes,
        )

    def test_chat_nao_gera_comunicado_e_nota_fica_vinculada_ao_atendimento(self):
        chamado = self.abrir()
        ChamadoService.assumir(chamado, self.atendente)
        ChamadoService.adicionar_mensagem(chamado, self.atendente, 'Análise iniciada.')
        ChamadoService.adicionar_mensagem(chamado, self.solicitante, 'Aguardando retorno.')

        comunicados = Comunicado.objects.filter(dados__chamado_id=chamado.pk)
        self.assertEqual(comunicados.count(), 1)
        self.assertEqual(comunicados.get().dados['evento_codigo'], 'ABERTURA')

        ChamadoService.resolver(
            chamado, self.atendente,
            causa_raiz='Configuração incorreta.', solucao='Configuração corrigida.',
        )
        self.assertEqual(comunicados.count(), 2)
        self.assertTrue(comunicados.filter(dados__evento_codigo='ENCERRAMENTO').exists())

        avaliacao = ChamadoService.avaliar(
            chamado, self.solicitante, nota=5, resolvido=True,
            comentario='Atendimento concluído.',
        )
        self.assertEqual(avaliacao.atendimento.atendente, self.atendente)
        self.assertEqual(avaliacao.atendimento.motivo_encerramento, 'RESOLVIDO')
        self.assertCountEqual(
            comunicados.values_list('dados__evento_codigo', flat=True),
            ['ABERTURA', 'ENCERRAMENTO', 'NOTA_ATENDIMENTO'],
        )
        nota = comunicados.get(dados__evento_codigo='NOTA_ATENDIMENTO')
        self.assertTrue(nota.usuarios.filter(pk=self.admin.pk).exists())
        self.assertFalse(nota.usuarios.filter(pk=self.solicitante.pk).exists())
        self.assertFalse(nota.usuarios.filter(pk=self.atendente.pk).exists())
        self.client.force_login(self.solicitante)
        resposta = self.client.get(reverse('chamados:detalhe', args=[chamado.pk]))
        self.assertEqual(resposta.context['comunicados_nao_lidos'], 2)
        self.assertNotContains(resposta, 'ATENDIMENTO CONCLUÍDO')

    def test_usuario_comum_nao_assume_nem_cria_nota_interna(self):
        chamado = self.abrir()
        with self.assertRaises(PermissionDenied):
            ChamadoService.assumir(chamado, self.solicitante)
        with self.assertRaises(ValidationError):
            ChamadoService.adicionar_mensagem(
                chamado, self.solicitante, 'Mensagem antes do aceite.'
            )
        ChamadoService.assumir(chamado, self.atendente)
        with self.assertRaises(PermissionDenied):
            ChamadoService.adicionar_mensagem(
                chamado, self.solicitante, 'Nota indevida', nota_interna=True
            )

    def test_exe_disfarcado_e_bloqueado(self):
        chamado = self.abrir()
        arquivo = SimpleUploadedFile('programa.exe', b'conteudo', content_type='application/octet-stream')

        with self.assertRaises(ValidationError):
            ChamadoService.adicionar_mensagem(
                chamado, self.solicitante, 'Segue o arquivo.', anexo=arquivo
            )

        self.assertFalse(ChamadoAnexo.objects.exists())

    def test_exe_e_rar_validos_sao_armazenados_com_nome_aleatorio(self):
        chamado = self.abrir()
        ChamadoService.assumir(chamado, self.atendente)
        exe = SimpleUploadedFile(
            'ferramenta.exe', b'MZ' + b'\0' * 32,
            content_type='application/octet-stream',
        )
        ChamadoService.adicionar_mensagem(
            chamado, self.atendente, 'Ferramenta para análise.', anexo=exe,
        )
        anexo = ChamadoAnexo.objects.get()
        self.assertEqual(anexo.nome_original, 'ferramenta.exe')
        self.assertNotIn('ferramenta', anexo.arquivo.name)
        rar = SimpleUploadedFile(
            'evidencias.rar', b'Rar!\x1a\x07\x00' + b'\0' * 32,
            content_type='application/vnd.rar',
        )
        ChamadoService.adicionar_mensagem(
            chamado, self.atendente, 'Evidências compactadas.', anexo=rar,
        )
        self.assertEqual(ChamadoAnexo.objects.count(), 2)

    def test_validadores_rar_leem_so_o_cabecalho_e_restauram_o_cursor(self):
        class UploadRastreado(BytesIO):
            name = 'evidencias.rar'
            content_type = 'application/octet-stream'

            def __init__(self, conteudo):
                super().__init__(conteudo)
                self.tamanhos_lidos = []

            @property
            def size(self):
                return len(self.getbuffer())

            def read(self, tamanho=-1):
                self.tamanhos_lidos.append(tamanho)
                return super().read(tamanho)

        arquivo = UploadRastreado(b'xxxRar!\x1a\x07\x00' + b'conteudo')
        arquivo.seek(3)

        validar_tamanho_anexo(arquivo)
        validar_mime_anexo(arquivo)
        validar_assinatura_anexo(arquivo)

        self.assertEqual(arquivo.tamanhos_lidos, [8])
        self.assertEqual(arquivo.tell(), 3)

    def test_rar_invalido_e_arquivo_acima_de_50_mb_sao_rejeitados_sem_copia(self):
        rar_invalido = SimpleUploadedFile(
            'renomeado.rar', b'isto nao e rar', content_type='application/octet-stream',
        )
        with self.assertRaisesMessage(
            ValidationError,
            'O CONTEÚDO NÃO CORRESPONDE A UM ARQUIVO RAR VÁLIDO.',
        ):
            validar_assinatura_anexo(rar_invalido)
        self.assertEqual(rar_invalido.tell(), 0)

        class ArquivoGrande:
            size = 50 * 1024 * 1024 + 1

        with self.assertRaisesMessage(
            ValidationError,
            'O ANEXO NÃO PODE ULTRAPASSAR 50 MB.',
        ):
            validar_tamanho_anexo(ArquivoGrande())

    @override_settings(
        FILE_UPLOAD_MAX_MEMORY_SIZE=1024 * 1024,
        FILE_UPLOAD_TEMP_DIR=gettempdir(),
    )
    def test_multipart_respeita_limiar_de_memoria_do_django(self):
        chamado = self.abrir()
        ChamadoService.assumir(chamado, self.atendente)
        self.client.force_login(self.atendente)
        url = reverse('chamados:mensagem', args=[chamado.pk])

        pequeno = SimpleUploadedFile(
            'pequeno.rar', b'Rar!\x1a\x07\x00' + b'x' * (100 * 1024),
            content_type='application/octet-stream',
        )
        with patch('chamados.views.ChamadoService.adicionar_mensagem') as adicionar:
            resposta = self.client.post(url, {'texto': 'Arquivo pequeno', 'anexo': pequeno})
        self.assertEqual(resposta.status_code, 302)
        self.assertIsInstance(adicionar.call_args.kwargs['anexo'], InMemoryUploadedFile)

        grande = SimpleUploadedFile(
            'grande.rar', b'Rar!\x1a\x07\x00' + b'x' * (1024 * 1024 + 1),
            content_type='application/octet-stream',
        )
        with patch('chamados.views.ChamadoService.adicionar_mensagem') as adicionar:
            resposta = self.client.post(url, {'texto': 'Arquivo grande', 'anexo': grande})
        self.assertEqual(resposta.status_code, 302)
        self.assertIsInstance(adicionar.call_args.kwargs['anexo'], TemporaryUploadedFile)

    def test_rotas_de_lista_dashboard_e_exportacao_respeitam_perfis(self):
        from openpyxl import load_workbook

        chamado = self.abrir()
        self.client.force_login(self.solicitante)
        self.assertEqual(self.client.get(reverse('chamados:lista')).status_code, 200)
        self.assertEqual(self.client.get(reverse('chamados:dashboard')).status_code, 403)
        self.assertEqual(self.client.get(reverse('chamados:exportar')).status_code, 403)

        self.client.force_login(self.admin)
        response = self.client.get(reverse('chamados:exportar'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        arquivo = BytesIO(b''.join(response.streaming_content))
        workbook = load_workbook(arquivo, read_only=True)
        planilha = workbook.active
        cabecalho, linha = planilha.iter_rows(values_only=True)
        self.assertEqual(cabecalho[2:4], ('SIGLA DA LOJA', 'NÚMERO DA LOJA'))
        self.assertEqual(linha[2:4], (self.cliente.sigla, chamado.loja))
        workbook.close()

    def test_dashboard_usa_tipo_do_equipamento_quando_categoria_de_suporte_esta_vazia(self):
        chamado = self.abrir()
        chamado.categoria = None
        chamado.save(update_fields=['categoria'])
        self.client.force_login(self.admin)

        response = self.client.get(
            reverse('chamados:dashboard'),
            {'periodo': '30d'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '"tipo_suporte": "Routers"')

    def test_solicitante_fecha_chamado_resolvido(self):
        chamado = self.abrir()
        ChamadoService.assumir(chamado, self.admin)
        ChamadoService.resolver(
            chamado, self.admin,
            causa_raiz='Credencial expirada.', solucao='Acesso normalizado.',
        )

        ChamadoService.avaliar(
            chamado, self.solicitante, nota=5, resolvido=True,
            comentario='Solução confirmada.',
        )
        chamado.refresh_from_db()
        self.assertEqual(chamado.status, Chamado.Status.ENCERRADO)

    def test_protocolos_sao_unicos_entre_empresas(self):
        primeiro = self.abrir()
        outra_empresa = Empresa.objects.create(nome='Outra Empresa')
        base = Base.objects.create(empresa=outra_empresa, nome='Base Matriz')
        cliente = Cliente.objects.create(sigla='EXT', nome='Cliente Externo')
        inventario = Inventario.objects.create(
            cliente=cliente, loja='Matriz', base=base,
            data_inicio=timezone.localdate(), criado_por=self.admin,
        )
        solicitante_segunda_empresa = User.objects.create_user(
            'solicitante_segunda_empresa', password='SenhaForte123!'
        )
        solicitante_segunda_empresa.perfil.empresa = outra_empresa
        solicitante_segunda_empresa.perfil.role = Perfil.Role.OPERADOR
        solicitante_segunda_empresa.perfil.save()
        solicitante_segunda_empresa.perfil.regionais.add(base)
        inventario.lider_usuario = solicitante_segunda_empresa
        inventario.save(update_fields=['lider_usuario'])

        segundo = ChamadoService.abrir(
            usuario=solicitante_segunda_empresa,
            base=base,
            inventario=inventario,
            equipamento=None,
            categoria=self.categoria,
            loja='',
            lider='',
            titulo='Falha geral',
            descricao='Teste de protocolo entre empresas.',
            prioridade=Chamado.Prioridade.NORMAL,
        )

        self.assertNotEqual(primeiro.protocolo, segundo.protocolo)

    def test_abertura_exige_inventario_e_operador_vinculado(self):
        with self.assertRaises(ValidationError):
            ChamadoService.abrir(
                usuario=self.solicitante, base=self.base, inventario=None,
                equipamento=None, categoria=self.categoria, loja='', lider='',
                titulo='Sem inventário', descricao='Não pode abrir.',
                prioridade=Chamado.Prioridade.NORMAL,
            )
        self.inventario.lider_usuario = self.outro
        self.inventario.save(update_fields=['lider_usuario'])
        with self.assertRaises(PermissionDenied):
            self.abrir()

    def test_aceite_e_sessoes_nao_permitam_duas_abertas(self):
        chamado = self.abrir()
        ChamadoService.assumir(chamado, self.atendente)
        chamado.refresh_from_db()
        self.assertIsNotNone(chamado.primeira_resposta_em)
        self.assertIsNotNone(chamado.aceito_em)
        self.assertEqual(chamado.sessoes.filter(encerrada_em__isnull=True).count(), 1)
        with self.assertRaises(ValidationError):
            ChamadoService._abrir_sessao(chamado, self.atendente, self.atendente)

        ChamadoService.alterar_status(
            chamado, self.atendente, Chamado.Status.AGUARDANDO_TERCEIRO,
            'Aguardando operadora.',
        )
        self.assertFalse(chamado.sessoes.filter(encerrada_em__isnull=True).exists())
        chamado.refresh_from_db()
        ChamadoService.alterar_status(
            chamado, self.atendente, Chamado.Status.EM_ATENDIMENTO,
        )
        self.assertEqual(chamado.sessoes.filter(encerrada_em__isnull=True).count(), 1)

    def test_avaliacao_negativa_reabre_e_positiva_encerra_depois_da_correcao(self):
        chamado = self.abrir()
        ChamadoService.assumir(chamado, self.atendente)
        ChamadoService.resolver(
            chamado, self.atendente, causa_raiz='Cabo solto', solucao='Cabo reconectado'
        )
        avaliacao = ChamadoService.avaliar(
            chamado, self.solicitante, nota=1, resolvido=False,
            comentario='Ainda sem rede.',
        )
        chamado.refresh_from_db()
        self.assertEqual(chamado.status, Chamado.Status.REABERTO)
        ChamadoService.alterar_status(chamado, self.atendente, Chamado.Status.EM_ATENDIMENTO)
        ChamadoService.resolver(
            chamado, self.atendente, causa_raiz='Porta queimada', solucao='Porta substituída'
        )
        segunda = ChamadoService.avaliar(
            chamado, self.solicitante, nota=5, resolvido=True,
            comentario='Resolvido.',
        )
        chamado.refresh_from_db()
        self.assertNotEqual(avaliacao.pk, segunda.pk)
        self.assertNotEqual(avaliacao.atendimento_id, segunda.atendimento_id)
        self.assertEqual(chamado.status, Chamado.Status.ENCERRADO)
        self.assertTrue(chamado.eventos.filter(tipo='REABERTURA_AVALIACAO').exists())

    def test_transferencia_fecha_sessao_e_preserva_atendentes(self):
        chamado = self.abrir()
        ChamadoService.assumir(chamado, self.atendente)
        ChamadoConexaoAtendente.objects.create(
            usuario=self.supervisor, canal='teste-transferencia-supervisor'
        )
        transferencia = ChamadoService.transferir_atendente(
            chamado, self.atendente, atendente_novo=self.supervisor,
            motivo='Escalonamento técnico.',
        )
        chamado.refresh_from_db()
        self.assertEqual(transferencia.atendente_anterior, self.atendente)
        self.assertEqual(chamado.atendente, self.supervisor)
        self.assertEqual(chamado.sessoes.filter(encerrada_em__isnull=True).get().atendente, self.supervisor)

    def test_conversao_sick_gera_sick_os_evento_e_comunicado(self):
        chamado = self.abrir()
        ChamadoService.assumir(chamado, self.atendente)
        sick = ChamadoService.converter_em_sick(
            chamado, self.atendente, diagnostico='Falha elétrica confirmada.'
        )
        chamado.refresh_from_db()
        self.assertEqual(chamado.sick, sick)
        self.assertTrue(Sick.objects.filter(pk=sick.pk, equipamento=self.equipamento).exists())
        self.assertTrue(OrdemServico.objects.filter(sick=sick).exists())
        self.assertTrue(chamado.eventos.filter(tipo='SICK_CRIADO').exists())

    def test_evento_e_imutavel(self):
        chamado = self.abrir()
        evento = chamado.eventos.first()
        evento.descricao = 'Alterado indevidamente'
        with self.assertRaises(ValidationError):
            evento.save()

    def test_alias_resolve_lider_e_nome_sem_alias_cria_pendencia(self):
        self.inventario.lider_usuario = None
        self.inventario.save(update_fields=['lider_usuario'])
        resultado = InventarioLiderService.resolver_texto_importado(self.inventario, self.admin)
        self.assertIsNone(resultado)
        self.assertTrue(PendenciaVinculoLider.objects.filter(inventario=self.inventario).exists())
        AliasUsuario.objects.create(usuario=self.solicitante, alias='Maria de Souza')
        self.inventario.lider = 'Mária de   Souza'
        self.inventario.save(update_fields=['lider'])
        resultado = InventarioLiderService.resolver_texto_importado(self.inventario, self.admin)
        self.inventario.refresh_from_db()
        self.assertEqual(resultado, self.solicitante)
        self.assertEqual(self.inventario.lider_usuario, self.solicitante)
        self.assertTrue(self.inventario.historico_vinculos_lider.exists())


class ChamadoWebSocketTests(TransactionTestCase):
    def setUp(self):
        self.empresa = Empresa.objects.create(nome='Empresa WebSocket')
        self.base = Base.objects.create(empresa=self.empresa, nome='Base WebSocket')
        self.outra_base = Base.objects.create(empresa=self.empresa, nome='Base Sem Acesso')
        self.usuario = User.objects.create_user('usuario_ws', password='SenhaForte123!')
        self.usuario.perfil.empresa = self.empresa
        self.usuario.perfil.role = Perfil.Role.OPERADOR
        self.usuario.perfil.save()
        self.usuario.perfil.regionais.add(self.base)
        suporte, _ = Group.objects.get_or_create(name=GruposChamados.SUPORTE)
        self.usuario.groups.add(suporte)
        self.intruso = User.objects.create_user('intruso_ws', password='SenhaForte123!')
        self.intruso.perfil.empresa = self.empresa
        self.intruso.perfil.role = Perfil.Role.OPERADOR
        self.intruso.perfil.save()
        self.intruso.perfil.regionais.add(self.outra_base)
        cliente = Cliente.objects.create(sigla='WSC', nome='Cliente WebSocket')
        inventario = Inventario.objects.create(
            cliente=cliente, loja='Loja WebSocket', base=self.base,
            data_inicio=timezone.localdate(), criado_por=self.usuario,
            lider='Usuário WS', lider_usuario=self.usuario,
        )
        categoria, _ = CategoriaChamado.objects.get_or_create(
            nome='OUTRO', defaults={'sla_horas': 24},
        )
        self.chamado = ChamadoService.abrir(
            usuario=self.usuario, base=self.base, inventario=inventario,
            equipamento=None, categoria=categoria, loja='Loja WebSocket',
            lider='Usuário WS', titulo='Teste de chat',
            descricao='Validar conversa em tempo real.',
            prioridade=Chamado.Prioridade.NORMAL,
        )

    def test_chat_e_vinculado_ao_chamado_e_respeita_acesso(self):
        ChamadoService.assumir(self.chamado, self.usuario)
        async def cenario():
            aplicacao = URLRouter(websocket_urlpatterns)
            autorizado = WebsocketCommunicator(
                aplicacao, f'/ws/chamados/{self.chamado.pk}/'
            )
            autorizado.scope['user'] = self.usuario
            conectado, _ = await autorizado.connect()
            self.assertTrue(conectado)
            await autorizado.send_json_to({'tipo': 'mensagem', 'texto': 'Olá pelo chat'})
            resposta = await autorizado.receive_json_from(timeout=3)
            self.assertEqual(resposta['tipo'], 'mensagem')
            self.assertEqual(resposta['item']['texto'], 'OLÁ PELO CHAT')
            await autorizado.disconnect()

            intruso = WebsocketCommunicator(
                aplicacao, f'/ws/chamados/{self.chamado.pk}/'
            )
            intruso.scope['user'] = self.intruso
            conectado, codigo = await intruso.connect()
            self.assertFalse(conectado)
            self.assertEqual(codigo, 4403)

        async_to_sync(cenario)()
        self.assertTrue(
            ChamadoMensagem.objects.filter(
                chamado=self.chamado, autor=self.usuario, texto='OLÁ PELO CHAT'
            ).exists()
        )

    def test_abertura_chega_a_todos_atendentes_conectados_da_base(self):
        suporte = Group.objects.get(name=GruposChamados.SUPORTE)
        segundo = User.objects.create_user('segundo_ws', password='SenhaForte123!')
        segundo.perfil.empresa = self.empresa
        segundo.perfil.save()
        segundo.perfil.regionais.add(self.base)
        segundo.groups.add(suporte)
        self.intruso.groups.add(suporte)

        async def cenario():
            aplicacao = URLRouter(websocket_urlpatterns)
            primeiro_ws = WebsocketCommunicator(aplicacao, '/ws/chamados/presenca/')
            primeiro_ws.scope['user'] = self.usuario
            segundo_ws = WebsocketCommunicator(aplicacao, '/ws/chamados/presenca/')
            segundo_ws.scope['user'] = segundo
            outra_base_ws = WebsocketCommunicator(aplicacao, '/ws/chamados/presenca/')
            outra_base_ws.scope['user'] = self.intruso

            for comunicador in (primeiro_ws, segundo_ws, outra_base_ws):
                conectado, _ = await comunicador.connect()
                self.assertTrue(conectado)

            await database_sync_to_async(ChamadoService._evento)(
                self.chamado,
                'ABERTURA',
                'CHAMADO ABERTO.',
                self.intruso,
            )

            primeiro_evento = await primeiro_ws.receive_json_from(timeout=3)
            segundo_evento = await segundo_ws.receive_json_from(timeout=3)
            for recebido in (primeiro_evento, segundo_evento):
                self.assertEqual(recebido['tipo'], 'chamado_evento')
                self.assertEqual(recebido['evento']['tipo'], 'ABERTURA')
                self.assertEqual(recebido['evento']['titulo'], self.chamado.titulo)
                self.assertEqual(recebido['evento']['base'], self.base.nome)
                self.assertEqual(
                    recebido['evento']['prioridade'],
                    self.chamado.get_prioridade_display(),
                )
            self.assertTrue(await outra_base_ws.receive_nothing(timeout=0.3))

            for comunicador in (primeiro_ws, segundo_ws, outra_base_ws):
                await comunicador.disconnect()

        async_to_sync(cenario)()
