from django.db.models.functions import TruncMonth
from django.db.models import (Q, Sum, F)
from django.http import HttpResponse
from insumos.models import ConsumoInsumo
from insumos.models import (
    AlteracaoCalendario,
    ChecklistDiario,
    Insumo,
    Inventario,
    MovimentacaoInsumo,
    SolicitacaoInsumo,
)
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from insumos.models import Insumo
from insumos.forms import (InsumoForm, CadastroInsumoForm)
from insumos.services.movimentacao_service import MovimentacaoService
from decimal import Decimal
from django.http import JsonResponse
from estoque.models import Equipamento, Produto
from insumos.models import LoteTag, RoloTag
from insumos.forms import FiltroEstoqueInsumoForm, InventarioForm
from django.db import transaction
from django.utils import timezone
from estoque.decorators import role_required
import openpyxl
import unicodedata
from django.contrib.admin.views.decorators import staff_member_required
from django.core.paginator import Paginator
from django.utils.dateparse import parse_date
from django.utils.timezone import make_aware
from datetime import datetime, date, time
from insumos.models import Inventario, Cliente
from estoque.models import Base, Empresa
from estoque.models import Comunicado
from django.contrib.auth.models import User
from insumos.services.checklist_service import ChecklistService
from insumos.services.movimentacao_service import MovimentacaoService
from collections import defaultdict
from io import BytesIO
from pathlib import Path
from django.conf import settings

@login_required
@role_required('admin', 'gestor', 'operador')
def estoque_insumos(request):
    perfil = request.user.perfil

    empresa_id = (request.GET.get('empresa') or '').strip()
    base_id = (request.GET.get('base') or '').strip()

    # =====================================================
    # ESCOPO DE BASES PERMITIDAS PARA O USUÁRIO
    # =====================================================
    if perfil.pode_ver_empresas_globais:
        bases_disponiveis_qs = (
            Base.objects
            .select_related('empresa')
            .all()
            .order_by('empresa__nome', 'nome')
        )
    else:
        bases_disponiveis_qs = (
            perfil.regionais
            .select_related('empresa')
            .all()
            .order_by('empresa__nome', 'nome')
        )

    bases_disponiveis = list(bases_disponiveis_qs)
    total_bases_disponiveis = len(bases_disponiveis)

    exibir_filtros = (
        perfil.pode_ver_empresas_globais
        or total_bases_disponiveis > 1
    )

    # =====================================================
    # EMPRESAS QUE PODEM SER EXIBIDAS NO FILTRO
    # =====================================================
    empresas = sorted(
        {
            base.empresa_id: base.empresa
            for base in bases_disponiveis
        }.values(),
        key=lambda empresa: empresa.nome,
    )

    # =====================================================
    # BASES UTILIZADAS NA CONSULTA
    # =====================================================
    bases_consulta = bases_disponiveis
    aguardando_filtro_base = False

    if exibir_filtros:
        if empresa_id:
            bases_consulta = [
                base for base in bases_consulta
                if str(base.empresa_id) == empresa_id
            ]

        if base_id:
            bases_consulta = [
                base for base in bases_consulta
                if str(base.id) == base_id
            ]
        else:
            # Evita materializar o estoque consolidado de dezenas de bases e
            # gerar uma página muito grande. A consulta operacional exige uma
            # base explícita quando o usuário possui mais de uma opção.
            bases_consulta = []
            aguardando_filtro_base = True

    else:
        # Com uma única base, ignora filtros enviados pela URL.
        empresa_id = ''
        base_id = ''

    # Uma única agregação substitui duas consultas para cada combinação
    # base/insumo. Somente pares com movimentação e saldo positivo são
    # materializados para o template.
    bases_por_id = {base.id: base for base in bases_consulta}
    saldos_agregados = list(
        MovimentacaoInsumo.objects
        .filter(
            base_id__in=bases_por_id,
            insumo__ativo=True,
        )
        .values('base_id', 'insumo_id')
        .annotate(
            entradas=Sum(
                'quantidade',
                filter=Q(tipo__in=[
                    'ENTRADA', 'DEVOLUCAO', 'AJUSTE_ENTRADA',
                ]),
            ),
            saidas=Sum(
                'quantidade',
                filter=Q(tipo__in=[
                    'SAIDA', 'PERDA', 'AJUSTE_SAIDA',
                ]),
            ),
        )
    )

    saldos_positivos = []
    insumos_ids = set()
    for agregado in saldos_agregados:
        saldo = (
            (agregado['entradas'] or Decimal('0')) -
            (agregado['saidas'] or Decimal('0'))
        )
        if saldo > 0:
            agregado['saldo'] = saldo
            saldos_positivos.append(agregado)
            insumos_ids.add(agregado['insumo_id'])

    insumos_por_id = {
        insumo.id: insumo
        for insumo in (
            Insumo.objects
            .filter(id__in=insumos_ids, ativo=True)
            .select_related('categoria')
        )
    }

    estoque = []
    for agregado in saldos_positivos:
        base = bases_por_id.get(agregado['base_id'])
        insumo = insumos_por_id.get(agregado['insumo_id'])
        if base is None or insumo is None:
            continue
        estoque.append({
            'base': base,
            'insumo': insumo,
            'saldo': agregado['saldo'],
            'minimo': insumo.estoque_minimo,
            'critico': agregado['saldo'] <= insumo.estoque_minimo,
        })

    estoque.sort(key=lambda item: (
        item['insumo'].categoria.nome,
        item['insumo'].descricao,
        item['base'].empresa.nome,
        item['base'].nome,
    ))

    # =====================================================
    # INDICADORES
    # =====================================================
    total_itens = len(estoque)
    criticos = sum(
        1
        for item in estoque
        if item['critico']
    )

    bases_com_saldo = len({
        item['base'].id
        for item in estoque
    })

    categorias_com_saldo = len({
        item['insumo'].categoria_id
        for item in estoque
    })

    # =====================================================
    # AGRUPAMENTO POR CATEGORIA
    # =====================================================
    estoque_por_categoria = defaultdict(list)

    for item in estoque:
        categoria_nome = item['insumo'].categoria.nome
        estoque_por_categoria[categoria_nome].append(item)

    return render(
        request,
        'insumos/estoque_insumos.html',
        {
            'estoque': estoque,
            'estoque_por_categoria': dict(
                estoque_por_categoria
            ),

            'total_itens': total_itens,
            'criticos': criticos,
            'ok_count': total_itens - criticos,
            'bases_com_saldo': bases_com_saldo,
            'categorias_com_saldo': categorias_com_saldo,

            # Opções completas dos filtros
            'empresas': empresas,
            'bases': bases_disponiveis,

            # Bases realmente consultadas
            'bases_consulta': bases_consulta,

            'exibir_filtros': exibir_filtros,
            'total_bases_disponiveis': total_bases_disponiveis,

            'filtro_empresa_id': empresa_id,
            'filtro_base_id': base_id,
            'aguardando_filtro_base': aguardando_filtro_base,
        },
    )

@login_required
def kpi_inventarios(request):

    perfil = request.user.perfil

    qs = Inventario.objects.all()

    if not perfil.pode_ver_empresas_globais:
        qs = qs.filter(base__in=perfil.regionais.all())

    data = {
        "planejados": qs.filter(status="PLANEJADO").count(),
        "andamento": qs.filter(status="EM_ANDAMENTO").count(),
        "finalizados": qs.filter(status="FINALIZADO").count(),
    }

    return JsonResponse(data)

@login_required
def consumo_por_base(request):

    data = (
        ConsumoInsumo.objects
        .values("inventario__base__nome")
        .annotate(total=Sum("valor_total"))
        .order_by("-total")
    )

    return JsonResponse(list(data), safe=False)

@login_required
def ranking_insumos(request):

    data = (
        ConsumoInsumo.objects
        .values("insumo__descricao")
        .annotate(total=Sum("quantidade"))
        .order_by("-total")[:10]
    )

    return JsonResponse(list(data), safe=False)

@login_required
def consumo_por_mes(request):

    data = (
        ConsumoInsumo.objects
        .annotate(mes=TruncMonth("criado_em"))
        .values("mes")
        .annotate(total=Sum("valor_total"))
        .order_by("mes")
    )

    return JsonResponse(list(data), safe=False)

@login_required
def lista_insumos(request):

    insumos = Insumo.objects.select_related('categoria').order_by(
        'categoria__nome',
        'descricao'
    )

    categoria = request.GET.get('categoria')

    if categoria:
        insumos = insumos.filter(categoria_id=categoria)

    return render(request, 'insumos/lista_insumos.html', {
        'insumos': insumos
    })

@login_required
def cadastrar_insumo(request):

    if request.method == 'POST':

        form = CadastroInsumoForm(request.POST, user=request.user)

        if form.is_valid():

            base = form.cleaned_data['base']
            insumo = form.cleaned_data['insumo']
            quantidade = form.cleaned_data['quantidade']

            MovimentacaoService.entrada(
                base=base,
                insumo=insumo,
                quantidade=quantidade,
                usuario=request.user,
                valor_unitario=Decimal('0.00'),
                observacao='Cadastro inicial de estoque'
            )

            messages.success(request, 'Entrada registrada com sucesso.')

            return redirect('insumos:cadastrar_insumos')

    else:

        form = CadastroInsumoForm(user=request.user)

    return render(request, 'insumos/cadastrar_insumos.html', {'form': form})

def get_equipamentos_disponiveis(request, categoria):

    regionais_ids = request.user.perfil.bases_checklist_ids

    if request.user.perfil.is_admin:
        equipamentos = Equipamento.objects.filter(
            status='ATIVO', finalidade=Equipamento.Finalidade.OPERACIONAL,
            produto__categoria=categoria,
        )
    else:
        equipamentos = Equipamento.objects.filter(
            status='ATIVO', finalidade=Equipamento.Finalidade.OPERACIONAL,
            produto__categoria=categoria, regional_id__in=regionais_ids,
        )

    data = [{
        'id': eq.id,
        'text': f"{eq.numero_serie} - {eq.produto.descricao} - {eq.patrimonio}",
        'numero_serie': eq.numero_serie,
        'patrimonio': eq.patrimonio
    } for eq in equipamentos]

    return JsonResponse({'results': data})

def get_lotes_tags_disponiveis(request):

    regionais_ids = request.user.perfil.bases_checklist_ids

    if request.user.perfil.is_admin:
        lotes = LoteTag.objects.filter(ativo=True, quantidade_disponivel__gt=0)
    else:
        lotes = LoteTag.objects.filter(ativo=True, quantidade_disponivel__gt=0, base_id__in=regionais_ids)

    data = [{
        'id': lote.id,
        'text': f"{lote.base.nome} - Lote {lote.numero_inicial} a {lote.numero_final} (Disponíveis: {lote.quantidade_disponivel})",
        'numero_inicial': lote.numero_inicial,
        'numero_final': lote.numero_final,
        'quantidade_disponivel': lote.quantidade_disponivel
    } for lote in lotes]

    return JsonResponse({'results': data})

@login_required
def editar_insumo(request, pk):

    insumo = get_object_or_404(Insumo, pk=pk)

    if request.method == 'POST':
        form = InsumoForm(request.POST, instance=insumo)

        if form.is_valid():
            form.save()
            messages.success(request, 'Insumo atualizado com sucesso.')
            return redirect('insumos:lista_insumos')

    else:
        form = InsumoForm(instance=insumo)

    return render(request,'insumos/cadastrar_insumo.html',
        {
            'form': form,
            'insumo': insumo
        }
    )

@login_required
def insumos_por_categoria(request):

    categoria_id = request.GET.get('categoria')

    if not categoria_id:
        return JsonResponse({'insumos': []})

    insumos = (Insumo.objects.filter(categoria_id=categoria_id, ativo=True).order_by('descricao').values('id', 'descricao'))

    return JsonResponse({'insumos': list(insumos)})

def finalizar_checklist_legado(checklist_id, usuario):
    checklist = ChecklistDiario.objects.get(id=checklist_id)

    with transaction.atomic():
        ChecklistService.finalizar(checklist=checklist, usuario=usuario)

def checklist_list(request):
    checklists = ChecklistDiario.objects.all()
    return render(request, 'insumos/checklist_list.html', {'checklists': checklists})

def serializar_valor(valor):
    """Converte objetos time, datetime, date para string serializável."""
    if isinstance(valor, time):
        return valor.strftime('%H:%M:%S')
    elif isinstance(valor, (datetime, date)):
        return valor.isoformat()
    elif isinstance(valor, Decimal):
        return float(valor)
    return valor

def normalizar_nome_base(valor):
    valor = unicodedata.normalize('NFKD', str(valor or ''))
    valor = ''.join(ch for ch in valor if not unicodedata.combining(ch))
    return ' '.join(valor.upper().split())

def regional_termina_com_x(valor):
    partes = normalizar_nome_base(valor).split()
    return bool(partes and partes[-1] == 'X')

def normalizar_cabecalho_excel(valor):
    valor = normalizar_nome_base(valor)
    return valor.replace('º', '').replace('°', '')

def encontrar_linha_cabecalho_calendario(sheet):
    max_linhas = min(sheet.max_row, 30)
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_linhas, values_only=True), start=1):
        valores = [normalizar_cabecalho_excel(valor) for valor in row if valor is not None]
        if (
            'SIGLA' in valores and
            'DATA' in valores and
            'REGIONAL' in valores and
            any('LOJA' in valor for valor in valores)
        ):
            return row_idx

    return None

def resolver_nome_base_importada(regional_nome, regional_map):
    nome_base = regional_map.get(regional_nome)
    if nome_base is None:
        regional_normalizado = normalizar_nome_base(regional_nome)
        mapa_normalizado = {
            normalizar_nome_base(chave): valor
            for chave, valor in regional_map.items()
        }
        nome_base = mapa_normalizado.get(regional_normalizado, regional_nome)

    nome_normalizado = normalizar_nome_base(nome_base)
    if not nome_normalizado or nome_normalizado == 'TODAS':
        return None

    if regional_termina_com_x(nome_base) and not normalizar_nome_base(nome_base).startswith('OXXO '):
        nome_oxxo = f'OXXO {str(nome_base).strip()}'
        nome_oxxo_normalizado = normalizar_nome_base(nome_oxxo)

        base_existente = next(
            (
                base
                for base in Base.objects.filter(empresa__nome__iexact='OXXO')
                if normalizar_nome_base(base.nome) == nome_oxxo_normalizado
            ),
            None
        )
        return base_existente.nome if base_existente else nome_oxxo

    if nome_normalizado in {'RJ', 'RIO DE JANEIRO'}:
        return 'RIO DE JANEIRO'

    if nome_normalizado.startswith('SP INT '):
        return nome_base

    if nome_normalizado.startswith('SP '):
        return 'SÃO PAULO'

    if 'CURITIBA' in nome_normalizado or 'FLORIPA' in nome_normalizado:
        return 'PR CURITIBA'

    return nome_base

def obter_base_importada(regional_nome, regional_map, empresa_padrao):
    nome_base = resolver_nome_base_importada(regional_nome, regional_map)
    if nome_base is None:
        return None

    empresa_base = empresa_para_base_importada(nome_base, empresa_padrao)
    nome_normalizado = normalizar_nome_base(nome_base)
    return next(
        (
            base
            for base in Base.objects.filter(empresa=empresa_base)
            if normalizar_nome_base(base.nome) == nome_normalizado
        ),
        None,
    )

def empresa_para_base_importada(nome_base, empresa_padrao):
    nome_normalizado = normalizar_nome_base(nome_base)
    if nome_normalizado.startswith('OXXO ') or regional_termina_com_x(nome_base):
        empresa_oxxo = Empresa.objects.filter(nome__iexact='OXXO').first()
        if empresa_oxxo:
            return empresa_oxxo

    return empresa_padrao

def texto_excel(valor):
    if valor is None:
        return ''
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()

def inteiro_excel(valor):
    texto = texto_excel(valor)
    if not texto:
        return None
    try:
        return int(float(texto.replace(',', '.')))
    except (TypeError, ValueError):
        return None

def data_excel_para_date(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor

    texto = texto_excel(valor)
    if not texto:
        return None

    data_parseada = parse_date(texto[:10])
    if data_parseada:
        return data_parseada

    for formato in ('%d/%m/%Y', '%d/%m/%y'):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue

    return None

def adicionar_aviso_importacao(resumo, mensagem, limite=50):
    resumo['avisos_total'] = resumo.get('avisos_total', 0) + 1
    if len(resumo['avisos']) < limite:
        resumo['avisos'].append(mensagem)

def importar_alteracoes_calendario(wb, arquivo_nome, usuario, regional_map, empresa_padrao, resumo):
    abas_alteracoes = [
        nome
        for nome in wb.sheetnames
        if normalizar_nome_base(nome) == 'ALTERACOES'
    ]
    if not abas_alteracoes:
        return 0

    total = 0
    blocos = [
        ('ATUAL', range(2, 10)),
        ('HISTORICO', range(11, 19)),
    ]

    for aba_nome in abas_alteracoes:
        sheet = wb[aba_nome]
        for origem_bloco, colunas in blocos:
            for row_idx in range(3, sheet.max_row + 1):
                valores = [sheet.cell(row=row_idx, column=col).value for col in colunas]
                if not any(valores):
                    continue

                revisao, data_valor, cliente_sigla, loja, descricao, regional_nome, solicitante, observacao = valores
                cliente_sigla = texto_excel(cliente_sigla)
                loja = texto_excel(loja)
                descricao = texto_excel(descricao)
                regional_nome = texto_excel(regional_nome)

                if not any([cliente_sigla, loja, descricao, regional_nome]):
                    continue

                data_alteracao = data_excel_para_date(data_valor)
                cliente = Cliente.objects.filter(sigla=cliente_sigla).first() if cliente_sigla else None
                base = None

                if regional_nome:
                    base = obter_base_importada(
                        regional_nome,
                        regional_map,
                        empresa_padrao,
                    )
                    if base is None and normalizar_nome_base(regional_nome) != 'TODAS':
                        adicionar_aviso_importacao(
                            resumo,
                            f'Aba Alteracoes, linha {row_idx}: regional '
                            f'{regional_nome} não corresponde a uma base cadastrada.'
                        )

                AlteracaoCalendario.objects.update_or_create(
                    origem_bloco=origem_bloco,
                    revisao=inteiro_excel(revisao),
                    data=data_alteracao,
                    cliente_sigla=cliente_sigla,
                    loja=loja,
                    descricao=descricao,
                    regional_nome=regional_nome,
                    defaults={
                        'cliente': cliente,
                        'base': base,
                        'solicitante': texto_excel(solicitante),
                        'observacao': texto_excel(observacao),
                        'arquivo': arquivo_nome,
                        'importado_por': usuario,
                    }
                )
                total += 1

                if not cliente and cliente_sigla:
                    adicionar_aviso_importacao(
                        resumo,
                        f'Aba Alteracoes, linha {row_idx}: cliente {cliente_sigla} nao encontrado.'
                    )

    return total

@staff_member_required
def importar_excel(request):
    if request.method == 'POST' and request.FILES.get('arquivo'):
        arquivo = request.FILES['arquivo']
        try:
            wb = openpyxl.load_workbook(arquivo, data_only=True)
        except Exception as e:
            messages.error(request, f'Erro ao ler o arquivo: {e}')
            return redirect('insumos:importar_excel')

        REGIONAL_MAP = {
            'PORTO ALEGRE': 'PORTO ALEGRE',
            'PR CURITIBA': 'PR CURITIBA',
            'PR MARINGÁ': 'PR MARINGÁ',
            'PR LONDRINA': 'PR LONDRINA',
            'PR PARANAGUÁ': 'PR PARANAGUÁ',
            'SC FLORIPA': 'SC FLORIPA',
            'SP INT CPN': 'SP INT CPN',
            'SP INT LIMEIRA': 'SP INT LIMEIRA',
            'SP INT BAURU': 'SP INT BAURU',
            'SP INT RIBEIRÃO': 'SP INT RIBEIRÃO',
            'SP INT STA ISA': 'SP INT STA ISABEL',
            'SP LESTE ITAQUA': 'SP LESTE ITAQUA',
            'SP LESTE': 'SÃO PAULO',
            'SP SUL': 'SÃO PAULO',
            'RJ': 'RIO DE JANEIRO',
            'SP LESTE X': 'OXXO SP LESTE X',
            'SP SUL X': 'OXXO SP SUL X',
            'SP LITORAL X': 'OXXO SP LITORAL X',
            'SP INT CPN X': 'OXXO SP INT CPN X',
            'SP INT JUNDIAÍ X': 'OXXO SP INT JUNDIAI X',
            'SP INT PIRACICABA X': 'OXXO SP INT PIRACICABA X',
            'SP INT SOROCABA X': 'OXXO SP INT SOROCABA X',
            'SP INT VALE X': 'OXXO SP INT VALE X',
            'SP LESTE GRU X': 'OXXO SP LESTE GRU X',
            'SP LESTE AND X': 'OXXO SP LESTE AND X',
        }

        # Mapeamento das colunas do Excel para os campos do modelo
        COLUMN_MAP = {
            'TIPO': 'tipo',
            'PESSOAS': 'pessoas',
            'OBSERVAÇÃO': 'observacao',
            'LÍDER': 'lider',
            'PONTO DE ENCONTRO': 'ponto_encontro',
            'HORÁRIO DO PONTO DE ENCONTRO': 'horario_ponto',
            'HORÁRIO DE INÍCIO': 'horario_inicio',
            'TIPO DA VISITA': 'tipo_visita',
            'RESPONSÁVEL PELA VISITA': 'responsavel_visita',
            'DATA DA VISITA': 'data_visita',
            'HORÁRIO DA VISITA': 'horario_visita',
            'RELATÓRIO DE VISITA': 'relatorio_visita',
            'PREP': 'prep',
            'HISTÓRICO EQUIPE': 'historico_equipe',
            'HISTÓRICO PEÇAS': 'historico_pecas',
            'HISTÓRICO SATISFAÇÃO': 'historico_satisfacao',
            'HISTÓRICO PREPARAÇÃO': 'historico_preparacao',
            'HISTÓRICO LÍDER': 'historico_lider',
            'HISTÓRICO DATA': 'historico_data',
            'EQUIPE PLAN': 'equipe_plan',
            'PREVISÃO DE PEÇAS': 'previsao_pecas',
            'PROD MÉDIA': 'prod_media',
            'BID': 'bid',
            'CNPJ': 'cnpj',
            'CEP': 'cep',
            'ENVIO DA ESCALA': 'envio_escala',
            'CHAVE': 'chave',
        }

        empresa_padrao = (
            Empresa.objects.filter(nome__iexact='Inventory Brasil').first()
            or Empresa.objects.first()
        )
        if not empresa_padrao:
            messages.error(request, 'Nenhuma empresa cadastrada.')
            return redirect('insumos:importar_excel')

        resumo_importacao = {
            'arquivo': arquivo.name,
            'clientes': 0,
            'alteracoes': 0,
            'inventarios_criados': 0,
            'inventarios_atualizados': 0,
            'bases_criadas': 0,
            'removidos': 0,
            'abas': [],
            'avisos': [],
            'avisos_total': 0,
        }

        abas_calendario = [
            (nome, encontrar_linha_cabecalho_calendario(wb[nome]))
            for nome in wb.sheetnames
        ]
        abas_calendario = [
            (nome, linha)
            for nome, linha in abas_calendario
            if linha
        ]

        with transaction.atomic():
            # 1. Processar aba "Siglas e Tipos" → Clientes
            if 'Siglas e Tipos' in wb.sheetnames:
                sheet = wb['Siglas e Tipos']
                clientes_processados = 0
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    sigla, nome, segmento, status = row[0], row[1], row[2], row[3]
                    if sigla and nome:
                        status_relatorio = str(status or '').strip().upper()
                        Cliente.objects.update_or_create(
                            sigla=sigla,
                            defaults={
                                'nome': nome,
                                'ativo': status_relatorio in {'ATIVO', 'LATAM'},
                                'status_relatorio': status_relatorio,
                            }
                        )
                        clientes_processados += 1
                resumo_importacao['clientes'] = clientes_processados

            resumo_importacao['alteracoes'] = importar_alteracoes_calendario(
                wb,
                arquivo.name,
                request.user,
                REGIONAL_MAP,
                empresa_padrao,
                resumo_importacao,
            )

            # 2. Processar abas de inventário
            if not abas_calendario:
                adicionar_aviso_importacao(resumo_importacao, 'Nenhuma aba de inventario encontrada.')
                request.session['resumo_importacao_excel'] = resumo_importacao
                messages.warning(request, 'Importacao finalizada sem abas de inventario. Confira o resumo abaixo.')
                return redirect('insumos:importar_excel')

            inventarios_importados_ids = set()
            escopos_importados = set()

            for aba_nome, cabecalho in abas_calendario:
                sheet = wb[aba_nome]

                # Encontrar linha de cabeçalho (SIGLA na coluna B)
                cabecalho = None
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if row and len(row) > 1 and row[1] == 'SIGLA':
                        cabecalho = row_idx
                        break

                if not encontrar_linha_cabecalho_calendario(sheet):
                    adicionar_aviso_importacao(resumo_importacao, f'Cabecalho nao encontrado na aba "{aba_nome}".')
                    continue

                cabecalho = encontrar_linha_cabecalho_calendario(sheet)
                header_row = list(sheet.iter_rows(min_row=cabecalho, max_row=cabecalho, values_only=True))[0]

                # Mapear índices das colunas obrigatórias
                col_map = {}
                for idx, col_name in enumerate(header_row):
                    if col_name == 'SIGLA':
                        col_map['sigla'] = idx
                    elif col_name == 'Nº DA LOJA':
                        col_map['loja'] = idx
                    elif col_name == 'DATA':
                        col_map['data'] = idx
                    elif col_name == 'ENDEREÇO':
                        col_map['endereco'] = idx
                    elif col_name == 'BAIRRO/NOME DA LOJA':
                        col_map['bairro'] = idx
                    elif col_name == 'CIDADE':
                        col_map['cidade'] = idx
                    elif col_name == 'REGIONAL':
                        col_map['regional'] = idx

                for idx, col_name in enumerate(header_row):
                    col_name_normalizado = normalizar_cabecalho_excel(col_name)
                    if col_name_normalizado == 'SIGLA':
                        col_map['sigla'] = idx
                    elif 'LOJA' in col_name_normalizado and 'BAIRRO' not in col_name_normalizado:
                        col_map['loja'] = idx
                    elif col_name_normalizado == 'DATA':
                        col_map['data'] = idx
                    elif col_name_normalizado == 'ENDERECO':
                        col_map['endereco'] = idx
                    elif 'BAIRRO' in col_name_normalizado:
                        col_map['bairro'] = idx
                    elif col_name_normalizado == 'CIDADE':
                        col_map['cidade'] = idx
                    elif col_name_normalizado == 'REGIONAL':
                        col_map['regional'] = idx

                colunas_obrigatorias = ['sigla', 'loja', 'data', 'regional']
                if not all(key in col_map for key in colunas_obrigatorias):
                    adicionar_aviso_importacao(resumo_importacao, f'Aba "{aba_nome}" nao possui colunas obrigatorias.')
                    continue

                contador = 0
                atualizados = 0
                for row_idx, row in enumerate(
                    sheet.iter_rows(min_row=cabecalho + 1, values_only=True),
                    start=cabecalho + 1,
                ):
                    if not row or not any(row):
                        continue

                    # --- Capturar campos obrigatórios ---
                    sigla = row[col_map['sigla']] if col_map.get('sigla') is not None and len(row) > col_map[
                        'sigla'] else None
                    loja = row[col_map['loja']] if col_map.get('loja') is not None and len(row) > col_map[
                        'loja'] else None
                    data_str = row[col_map['data']] if col_map.get('data') is not None and len(row) > col_map[
                        'data'] else None
                    endereco = row[col_map['endereco']] if col_map.get('endereco') is not None and len(row) > col_map[
                        'endereco'] else ''
                    bairro = row[col_map['bairro']] if col_map.get('bairro') is not None and len(row) > col_map[
                        'bairro'] else ''
                    cidade = row[col_map['cidade']] if col_map.get('cidade') is not None and len(row) > col_map[
                        'cidade'] else ''
                    regional_nome = row[col_map['regional']] if col_map.get('regional') is not None and len(row) > \
                                                                col_map['regional'] else None

                    if not sigla or not loja or not data_str or not regional_nome:
                        continue

                    # --- Capturar TODAS as colunas (incluindo as extras) ---
                    dados_completos = {}
                    for idx, valor in enumerate(row):
                        if idx < len(header_row):
                            nome_coluna = header_row[idx]
                        else:
                            nome_coluna = f"col_{idx}"
                        if valor is not None:
                            dados_completos[nome_coluna] = serializar_valor(valor)

                    # Remover colunas que já extraímos (para não duplicar)
                    colunas_extraidas = ['SIGLA', 'Nº DA LOJA', 'DATA', 'ENDEREÇO', 'BAIRRO/NOME DA LOJA', 'CIDADE',
                                         'REGIONAL']
                    for col in colunas_extraidas:
                        dados_completos.pop(col, None)

                    dados_completos['_importacao_calendario'] = {
                        'arquivo': arquivo.name,
                        'aba': aba_nome,
                        'linha': row_idx,
                        'regional_excel': regional_nome,
                        'importado_em': timezone.now().isoformat(),
                    }

                    # --- Buscar cliente ---
                    try:
                        cliente = Cliente.objects.get(sigla=sigla)
                    except Cliente.DoesNotExist:
                        adicionar_aviso_importacao(
                            resumo_importacao,
                            f'Aba {aba_nome}, linha {row_idx}: cliente {sigla} nao encontrado.'
                        )
                        continue

                    # --- Mapear regional ---
                    base = obter_base_importada(
                        regional_nome,
                        REGIONAL_MAP,
                        empresa_padrao,
                    )
                    if base is None:
                        adicionar_aviso_importacao(
                            resumo_importacao,
                            f'Aba {aba_nome}, linha {row_idx}: regional '
                            f'{regional_nome} não corresponde a uma base cadastrada.'
                        )
                        continue

                    # --- Converter data ---
                    if isinstance(data_str, datetime):
                        data_inicio = data_str.date()
                    else:
                        data_inicio = parse_date(str(data_str)) if data_str else None
                    if not data_inicio:
                        adicionar_aviso_importacao(
                            resumo_importacao,
                            f'Aba {aba_nome}, linha {row_idx}: data invalida para {sigla} loja {loja}.'
                        )
                        continue

                    # --- Preparar defaults com campos extras ---
                    defaults = {
                        'status': 'PLANEJADO',
                        'criado_por': request.user,
                        'endereco': endereco,
                        'bairro': bairro,
                        'cidade': cidade,
                        'dados_brutos': dados_completos,
                    }

                    # --- Mapear colunas extras para os campos do modelo ---
                    for excel_col, model_field in COLUMN_MAP.items():
                        valor = dados_completos.get(excel_col)
                        if valor is not None and str(valor).strip():
                            # Conversão de tipos
                            if model_field in ['pessoas', 'equipe_plan', 'previsao_pecas']:
                                try:
                                    valor = int(float(valor))
                                except (ValueError, TypeError):
                                    valor = None
                            elif model_field in ['prep', 'prod_media']:
                                try:
                                    valor = float(valor)
                                except (ValueError, TypeError):
                                    valor = None
                            elif model_field in ['horario_ponto', 'horario_inicio', 'horario_visita']:
                                if isinstance(valor, time):
                                    pass  # já é time
                                else:
                                    try:
                                        if isinstance(valor, str) and ':' in valor:
                                            valor = datetime.strptime(valor, '%H:%M:%S').time()
                                        else:
                                            valor = None
                                    except:
                                        valor = None
                            elif model_field in ['data_visita', 'historico_data', 'envio_escala']:
                                if isinstance(valor, datetime):
                                    valor = valor.date()
                                else:
                                    try:
                                        valor = parse_date(str(valor))
                                    except:
                                        valor = None
                            defaults[model_field] = valor

                    # O tipo faz parte da identidade: uma loja pode ter CA, CP e T no mesmo dia.
                    tipo_inventario = defaults.get('tipo') or ''
                    inventario, created = Inventario.objects.get_or_create(
                        cliente=cliente,
                        loja=str(loja),
                        base=base,
                        data_inicio=data_inicio,
                        tipo=tipo_inventario,
                        defaults=defaults
                    )
                    if not created:
                        from integracao.models import InventoryPlanningEventBinding
                        sincronizado_api = InventoryPlanningEventBinding.objects.filter(
                            inventory=inventario
                        ).exists()
                        if sincronizado_api:
                            adicionar_aviso_importacao(
                                resumo_importacao,
                                f'Aba {aba_nome}, linha {row_idx}: inventário vinculado '
                                'à Inventory Planning API não foi sobrescrito pela planilha.'
                            )
                        else:
                            for key, value in defaults.items():
                                setattr(inventario, key, value)
                            inventario.save()
                            atualizados += 1
                    else:
                        contador += 1

                    inventarios_importados_ids.add(inventario.id)
                    escopos_importados.add((data_inicio.year, data_inicio.month, base.id))

                resumo_importacao['inventarios_criados'] += contador
                resumo_importacao['inventarios_atualizados'] += atualizados
                resumo_importacao['abas'].append({
                    'nome': aba_nome,
                    'criados': contador,
                    'atualizados': atualizados,
                })


            removidos_revisao = 0
            for ano, mes, base_id in escopos_importados:
                obsoletos = (
                    Inventario.objects
                    .filter(
                        base_id=base_id,
                        data_inicio__year=ano,
                        data_inicio__month=mes,
                        status='PLANEJADO',
                        checklists__isnull=True,
                    )
                    .filter(planning_event_binding__isnull=True)
                    .exclude(id__in=inventarios_importados_ids)
                )
                quantidade = obsoletos.count()
                if quantidade:
                    obsoletos.delete()
                    removidos_revisao += quantidade

            if removidos_revisao:
                messages.info(
                    request,
                    f'{removidos_revisao} inventários planejados ausentes da revisão atual foram removidos.'
                )

            resumo_importacao['removidos'] = removidos_revisao
            request.session['resumo_importacao_excel'] = resumo_importacao
            messages.success(request, 'Importação concluída!')
            return redirect('insumos:importar_excel')

    resumo_importacao = request.session.pop('resumo_importacao_excel', None)
    return render(request, 'insumos/importar_excel.html', {'resumo_importacao': resumo_importacao})

@login_required
def inventario_detalhes(request, inventario_id):
    inventario = get_object_or_404(Inventario, pk=inventario_id)
    data = {
        'cliente': f"{inventario.cliente.sigla} - {inventario.cliente.nome}",
        'sigla': inventario.cliente.sigla,
        'loja': inventario.loja,
        'base': inventario.base.nome,
        'data': inventario.data_inicio.strftime('%d/%m/%Y'),
        'endereco': inventario.endereco or '',
        'bairro': inventario.bairro or '',
        'cidade': inventario.cidade or '',
        'lider': inventario.lider or '',
        'ponto_encontro': inventario.ponto_encontro or '',
        'horario_ponto': (
            inventario.horario_ponto.strftime('%H:%M')
            if inventario.horario_ponto else ''
        ),
        'horario_inicio': (
            inventario.horario_inicio.strftime('%H:%M')
            if inventario.horario_inicio else ''
        ),
        'pessoas': inventario.pessoas,
        'limite_coletores': (
            inventario.pessoas + 5
            if inventario.pessoas is not None
            else None
        ),
    }
    return JsonResponse(data)

def media_pecas_por_cliente(request):

    inventarios = Inventario.objects.filter(status='FINALIZADO')

    total = 0
    count = 0
    for inv in inventarios:
        previsao = inv.dados_brutos.get('PREVISÃO DE PEÇAS')
        if previsao and isinstance(previsao, (int, float)):
            total += previsao
            count += 1

    media = total / count if count > 0 else 0
    return JsonResponse({'media_pecas': media})

def planejamento_inventarios(request):
    inventarios = Inventario.objects.filter(status='PLANEJADO')
    dados = []
    for inv in inventarios:
        dados.append({
            'cliente': inv.cliente.sigla,
            'loja': inv.loja,
            'data': inv.data_inicio,
            'tipo': inv.dados_brutos.get('TIPO'),
            'pessoas': inv.dados_brutos.get('PESSOAS'),
            'lider': inv.dados_brutos.get('LÍDER'),
            'previsao_pecas': inv.dados_brutos.get('PREVISÃO DE PEÇAS'),
            'prod_media': inv.dados_brutos.get('PROD MÉDIA'),
            'observacao': inv.dados_brutos.get('OBSERVAÇÃO'),
        })
    return render(request, 'planejamento.html', {'inventarios': dados})

@login_required
def gerenciar_inventarios(request):
    perfil = request.user.perfil
    if not (perfil.is_admin or perfil.is_planejamento_insumos):
        messages.error(request, 'Acesso restrito a administradores e planejamento.')
        return redirect('estoque:index')

    # Filtros
    cliente_id = request.GET.get('cliente')
    regional_id = request.GET.get('regional')
    status_filter = request.GET.get('status')
    hoje = timezone.localdate().isoformat()
    data_dia = request.GET.get('data') or hoje
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    inventarios = (
        Inventario.objects
        .select_related('cliente', 'base')
        .only(
            'id', 'cliente__sigla', 'cliente__nome', 'base__nome', 'criado_por_id',
            'loja', 'data_inicio', 'status', 'endereco', 'bairro', 'cidade',
            'lider', 'ponto_encontro', 'previsao_pecas', 'bid', 'cnpj', 'chave',
        )
    )

    if perfil.pode_ver_empresas_globais:
        inventarios = inventarios.all()
    else:
        inventarios = inventarios.filter(base__in=perfil.regionais.all())

    if cliente_id:
        inventarios = inventarios.filter(cliente_id=cliente_id)
    if regional_id:
        inventarios = inventarios.filter(base_id=regional_id)
    if status_filter:
        inventarios = inventarios.filter(status=status_filter)
    if data_dia and not data_inicio and not data_fim:
        inventarios = inventarios.filter(data_inicio=data_dia)
    elif data_inicio:
        inventarios = inventarios.filter(data_inicio__gte=data_inicio)
    if data_fim:
        inventarios = inventarios.filter(data_inicio__lte=data_fim)

    inventarios = inventarios.order_by('-data_inicio', 'cliente__sigla', 'loja')
    total_inventarios = inventarios.count()
    paginator = Paginator(inventarios, 50)
    page_obj = paginator.get_page(request.GET.get('page'))
    query_params = request.GET.copy()
    query_params.pop('page', None)

    context = {
        'inventarios': page_obj.object_list,
        'page_obj': page_obj,
        'total_inventarios': total_inventarios,
        'querystring': query_params.urlencode(),
        'clientes': Cliente.objects.filter(ativo=True).order_by('sigla'),
        'regionais': (
            Base.objects.all().order_by('nome')
            if perfil.pode_ver_empresas_globais
            else perfil.regionais.all().order_by('nome')
        ),
        'status_choices': Inventario.STATUS,
        'filtro_cliente': cliente_id,
        'filtro_regional': regional_id,
        'filtro_status': status_filter,
        'filtro_data': data_dia,
    }
    return render(request, 'insumos/gerenciar_inventarios.html', context)

@login_required
def lista_inventarios(request):
    # Filtros via GET
    sigla = request.GET.get('sigla', '')
    loja = request.GET.get('loja', '')
    hoje = timezone.localdate().isoformat()
    data_dia = request.GET.get('data') or hoje
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    regional_id = request.GET.get('regional', '')

    # Base queryset com permissões do usuário
    if request.user.perfil.pode_ver_empresas_globais:
        inventarios = Inventario.objects.all().select_related('cliente', 'base')
    else:
        inventarios = Inventario.objects.filter(
            base__in=request.user.perfil.regionais.all()
        ).select_related('cliente', 'base')
    inventarios = inventarios.only(
        'id', 'cliente__sigla', 'cliente__nome', 'base__nome',
        'loja', 'data_inicio', 'status', 'pessoas', 'tipo',
        'lider', 'ponto_encontro', 'previsao_pecas', 'bid', 'cidade',
    )

    # Aplicar filtros
    if sigla:
        inventarios = inventarios.filter(cliente__sigla__icontains=sigla)
    if loja:
        inventarios = inventarios.filter(loja__icontains=loja)
    if data_dia and not data_inicio and not data_fim:
        inventarios = inventarios.filter(data_inicio=data_dia)
    elif data_inicio:
        inventarios = inventarios.filter(data_inicio__gte=data_inicio)
    if data_fim:
        inventarios = inventarios.filter(data_inicio__lte=data_fim)
    if regional_id:
        inventarios = inventarios.filter(base_id=regional_id)

    # Ordenar por data mais recente
    inventarios = inventarios.order_by('-data_inicio', 'cliente__sigla', 'loja')
    total_inventarios = inventarios.count()
    paginator = Paginator(inventarios, 50)
    page_obj = paginator.get_page(request.GET.get('page'))
    query_params = request.GET.copy()
    query_params.pop('page', None)

    # Lista de regionais para o filtro
    if request.user.perfil.pode_ver_empresas_globais:
        regionais = Base.objects.all().order_by('nome')
    else:
        regionais = request.user.perfil.regionais.all().order_by('nome')

    context = {
        'inventarios': page_obj.object_list,
        'page_obj': page_obj,
        'total_inventarios': total_inventarios,
        'querystring': query_params.urlencode(),
        'regionais': regionais,
        'perfil': request.user.perfil,
        'filtros': {
            'sigla': sigla,
            'loja': loja,
            'data': data_dia,
            'regional_id': regional_id,
        }
    }
    return render(request, 'insumos/lista_inventarios.html', context)

@login_required
def editar_inventario(request, pk):
    inventario = get_object_or_404(Inventario, pk=pk)
    perfil = request.user.perfil
    pode_editar = perfil.is_admin or perfil.is_gestor or perfil.is_planejamento_insumos
    if not pode_editar:
        messages.error(request, 'Você não tem permissão para editar.')
        return redirect('insumos:lista_inventarios')
    if not perfil.pode_ver_empresas_globais and not perfil.regionais.filter(id=inventario.base_id).exists():
        messages.error(request, 'Sem permissao para editar esta base.')
        return redirect('insumos:lista_inventarios')

    if request.method == 'POST':
        form = InventarioForm(request.POST, instance=inventario)
        if form.is_valid():
            # Atualiza campos principais
            inventario = form.save(commit=False)
            # Atualiza dados_brutos com os campos dinâmicos
            from integracao.models import InventoryPlanningEventBinding
            sincronizado = InventoryPlanningEventBinding.objects.filter(
                inventory=inventario
            ).exists()
            if not sincronizado:
                dados = {}
                for key, value in request.POST.items():
                    if key.startswith('campo_'):
                        nome_campo = key.replace('campo_', '')
                        dados[nome_campo] = value
                inventario.dados_brutos = dados
            inventario.save()
            messages.success(request, 'Inventário atualizado com sucesso.')
            return redirect('insumos:lista_inventarios')
    else:
        form = InventarioForm(instance=inventario)

    return render(request, 'insumos/editar_inventario.html', {
        'form': form,
        'inventario': inventario
    })

@login_required
def editar_inventario_modal(request, inventario_id):
    if not request.user.perfil.is_admin:
        return JsonResponse({'error': 'Acesso negado'}, status=403)

    inventario = get_object_or_404(Inventario, pk=inventario_id)
    if request.method == 'POST':
        form = InventarioForm(request.POST, instance=inventario)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True})
        return JsonResponse({'errors': form.errors}, status=400)

    from django.template.loader import render_to_string
    form = InventarioForm(instance=inventario)
    html = render_to_string('insumos/partials/editar_inventario.html', {'form': form, 'inventario': inventario})
    return JsonResponse({'html': html})

@login_required
def exportar_excel(request):
    """Exporta todos os inventários para um arquivo Excel no mesmo formato do importado."""

    # Criar workbook
    wb = openpyxl.Workbook()

    # --- ABA 1: Siglas e Tipos ---
    ws_siglas = wb.active
    ws_siglas.title = "Siglas e Tipos"

    # Cabeçalho
    cabecalho_siglas = ['SIGLA', 'CLIENTE', 'SEGMENTO', 'STATUS', '', '2026', 'TOTAL CLIENTES', '', '', 'TIPO',
                        'DESCRIÇÃO']
    ws_siglas.append(cabecalho_siglas)

    # Dados dos clientes ativos
    clientes = Cliente.objects.filter(ativo=True).order_by('sigla')
    for cliente in clientes:
        ws_siglas.append([
            cliente.sigla,
            cliente.nome,
            '',  # SEGMENTO (se não tiver, deixar em branco)
            'ATIVO',
            '',
            '',
            '',
            '',
            '',
            '',
            ''
        ])

    # --- ABA 2: Alterações (opcional, pode deixar vazia ou com cabeçalho) ---
    ws_alteracoes = wb.create_sheet("Alterações")
    ws_alteracoes.append(['', 'REGISTRO DE ALTERAÇÕES - ' + datetime.now().strftime('%B %Y').upper()])
    ws_alteracoes.append(
        ['', 'Revisão', 'Data', 'Cliente', 'Nº da Loja', 'Descrição da Alteração', 'Regional', 'Solicitante', 'Obs.'])
    # Deixar vazio ou adicionar dados se tiver histórico de alterações

    # --- ABA 3: Dados do mês atual (ex: JUN) ---
    mes_atual = datetime.now().strftime('%b').upper()  # Ex: "JUN"
    aba_principal = mes_atual

    # Verificar se a aba já existe (se criamos "Siglas e Tipos" e "Alterações", a próxima é a 3ª)
    # Como já temos 2 abas, a próxima será a terceira
    ws_principal = wb.create_sheet(aba_principal)

    # Cabeçalho da planilha principal (mesmo do arquivo importado)
    cabecalho_principal = [
        '', 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20,
        21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35
    ]
    ws_principal.append(cabecalho_principal)

    # Segunda linha: cabeçalho das colunas (igual ao Excel)
    cabecalho_colunas = [
        '#', 'SIGLA', 'Nº DA LOJA', 'DATA', 'ENDEREÇO', 'BAIRRO/NOME DA LOJA', 'CIDADE', 'REGIONAL',
        'TIPO', 'PESSOAS', 'OBSERVAÇÃO', 'LÍDER', 'PONTO DE ENCONTRO', 'HORÁRIO DO PONTO DE ENCONTRO',
        'HORÁRIO DE INÍCIO', 'TIPO DA VISITA', 'RESPONSÁVEL PELA VISITA', 'DATA DA VISITA',
        'HORÁRIO DA VISITA', 'RELATÓRIO DE VISITA', 'PREP', 'HISTÓRICO EQUIPE', 'HISTÓRICO PEÇAS',
        'HISTÓRICO SATISFAÇÃO', 'HISTÓRICO PREPARAÇÃO', 'HISTÓRICO LÍDER', 'HISTÓRICO DATA',
        'EQUIPE PLAN', 'PREVISÃO DE PEÇAS', 'PROD MÉDIA', 'BID', 'CNPJ', 'CEP', 'ENVIO DA ESCALA', 'CHAVE'
    ]
    ws_principal.append(cabecalho_colunas)

    # Dados dos inventários (apenas os que estão planejados ou em andamento, ou todos?)
    inventarios = Inventario.objects.filter(
        status__in=['PLANEJADO', 'EM_ANDAMENTO']
    ).select_related('cliente', 'base')

    # Ordenar por data e cliente
    inventarios = inventarios.order_by('data_inicio', 'cliente__sigla')

    # Para cada inventário, criar uma linha no Excel
    for idx, inv in enumerate(inventarios, start=1):
        # Usar os dados do banco ou, se existir, os dados brutos salvos
        dados = inv.dados_brutos if inv.dados_brutos else {}

        linha = [
            idx,  # #
            inv.cliente.sigla,
            inv.loja,
            inv.data_inicio.strftime('%Y-%m-%d') if inv.data_inicio else '',
            inv.endereco or '',
            inv.bairro or '',
            inv.cidade or '',
            inv.base.nome,
            inv.tipo or dados.get('TIPO', ''),
            inv.pessoas or dados.get('PESSOAS', ''),
            inv.observacao or dados.get('OBSERVAÇÃO', ''),
            inv.lider or dados.get('LÍDER', ''),
            inv.ponto_encontro or dados.get('PONTO DE ENCONTRO', ''),
            inv.horario_ponto.strftime('%H:%M:%S') if inv.horario_ponto else dados.get('HORÁRIO DO PONTO DE ENCONTRO',
                                                                                       ''),
            inv.horario_inicio.strftime('%H:%M:%S') if inv.horario_inicio else dados.get('HORÁRIO DE INÍCIO', ''),
            inv.tipo_visita or dados.get('TIPO DA VISITA', ''),
            inv.responsavel_visita or dados.get('RESPONSÁVEL PELA VISITA', ''),
            inv.data_visita.strftime('%Y-%m-%d') if inv.data_visita else dados.get('DATA DA VISITA', ''),
            inv.horario_visita.strftime('%H:%M:%S') if inv.horario_visita else dados.get('HORÁRIO DA VISITA', ''),
            inv.relatorio_visita or dados.get('RELATÓRIO DE VISITA', ''),
            inv.prep or dados.get('PREP', ''),
            inv.historico_equipe or dados.get('HISTÓRICO EQUIPE', ''),
            inv.historico_pecas or dados.get('HISTÓRICO PEÇAS', ''),
            inv.historico_satisfacao or dados.get('HISTÓRICO SATISFAÇÃO', ''),
            inv.historico_preparacao or dados.get('HISTÓRICO PREPARAÇÃO', ''),
            inv.historico_lider or dados.get('HISTÓRICO LÍDER', ''),
            inv.historico_data.strftime('%Y-%m-%d') if inv.historico_data else dados.get('HISTÓRICO DATA', ''),
            inv.equipe_plan or dados.get('EQUIPE PLAN', ''),
            inv.previsao_pecas or dados.get('PREVISÃO DE PEÇAS', ''),
            inv.prod_media or dados.get('PROD MÉDIA', ''),
            inv.bid or dados.get('BID', ''),
            inv.cnpj or dados.get('CNPJ', ''),
            inv.cep or dados.get('CEP', ''),
            inv.envio_escala.strftime('%Y-%m-%d') if inv.envio_escala else dados.get('ENVIO DA ESCALA', ''),
            inv.chave or dados.get('CHAVE', ''),
        ]

        # Garantir que a linha tenha exatamente 35 colunas
        while len(linha) < 35:
            linha.append('')
        ws_principal.append(linha)

    # --- ABA 4: Aba OXX (se houver inventários OXX) ---
    inventarios_oxx = inventarios.filter(cliente__sigla='OXX')
    if inventarios_oxx.exists():
        aba_oxx = mes_atual + " OXX"
        ws_oxx = wb.create_sheet(aba_oxx)

        # Cabeçalho igual ao da aba principal
        ws_oxx.append(cabecalho_principal)
        ws_oxx.append(cabecalho_colunas)

        for idx, inv in enumerate(inventarios_oxx, start=1):
            dados = inv.dados_brutos if inv.dados_brutos else {}
            linha = [
                idx,
                inv.cliente.sigla,
                inv.loja,
                inv.data_inicio.strftime('%Y-%m-%d') if inv.data_inicio else '',
                inv.endereco or '',
                inv.bairro or '',
                inv.cidade or '',
                inv.base.nome,
                inv.tipo or dados.get('TIPO', ''),
                inv.pessoas or dados.get('PESSOAS', ''),
                inv.observacao or dados.get('OBSERVAÇÃO', ''),
                inv.lider or dados.get('LÍDER', ''),
                inv.ponto_encontro or dados.get('PONTO DE ENCONTRO', ''),
                inv.horario_ponto.strftime('%H:%M:%S') if inv.horario_ponto else dados.get(
                    'HORÁRIO DO PONTO DE ENCONTRO', ''),
                inv.horario_inicio.strftime('%H:%M:%S') if inv.horario_inicio else dados.get('HORÁRIO DE INÍCIO', ''),
                inv.tipo_visita or dados.get('TIPO DA VISITA', ''),
                inv.responsavel_visita or dados.get('RESPONSÁVEL PELA VISITA', ''),
                inv.data_visita.strftime('%Y-%m-%d') if inv.data_visita else dados.get('DATA DA VISITA', ''),
                inv.horario_visita.strftime('%H:%M:%S') if inv.horario_visita else dados.get('HORÁRIO DA VISITA', ''),
                inv.relatorio_visita or dados.get('RELATÓRIO DE VISITA', ''),
                inv.prep or dados.get('PREP', ''),
                inv.historico_equipe or dados.get('HISTÓRICO EQUIPE', ''),
                inv.historico_pecas or dados.get('HISTÓRICO PEÇAS', ''),
                inv.historico_satisfacao or dados.get('HISTÓRICO SATISFAÇÃO', ''),
                inv.historico_preparacao or dados.get('HISTÓRICO PREPARAÇÃO', ''),
                inv.historico_lider or dados.get('HISTÓRICO LÍDER', ''),
                inv.historico_data.strftime('%Y-%m-%d') if inv.historico_data else dados.get('HISTÓRICO DATA', ''),
                inv.equipe_plan or dados.get('EQUIPE PLAN', ''),
                inv.previsao_pecas or dados.get('PREVISÃO DE PEÇAS', ''),
                inv.prod_media or dados.get('PROD MÉDIA', ''),
                inv.bid or dados.get('BID', ''),
                inv.cnpj or dados.get('CNPJ', ''),
                inv.cep or dados.get('CEP', ''),
                inv.envio_escala.strftime('%Y-%m-%d') if inv.envio_escala else dados.get('ENVIO DA ESCALA', ''),
                inv.chave or dados.get('CHAVE', ''),
            ]
            while len(linha) < 35:
                linha.append('')
            ws_oxx.append(linha)

    # --- Salvar arquivo ---
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nome_arquivo = f'inventarios_{datetime.now().strftime("%Y_%m_%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'

    wb.save(response)
    return response

@login_required
def insumos_por_base(request):
    base_id = request.GET.get('base_id')
    if not base_id:
        return JsonResponse({'insumos': []})

    perfil = request.user.perfil
    base = get_object_or_404(Base.objects.select_related('empresa'), pk=base_id)
    if (
        not perfil.is_admin and
        (
            base.pk not in perfil.bases_checklist_ids or
            base.empresa_id != perfil.empresa_id
        )
    ):
        return JsonResponse(
            {'insumos': [], 'erro': 'Base não autorizada.'},
            status=403,
        )

    data = [
        {
            'id': item['id'],
            'descricao': item['descricao'],
            'categoria': item['categoria'],
            'unidade': item['insumo'].unidade_medida,
            'saldo': float(item['saldo']),
        }
        for item in ChecklistService.insumos_disponiveis_para_checklist(base)
    ]
    return JsonResponse({'insumos': data})

@login_required
def lista_checklists(request):
    checklists = (
        ChecklistDiario.objects
        .select_related(
            "inventario",
            "inventario__cliente",
            "inventario__base",
            "criado_por",
        )
        .order_by("-data_inicio", "-id")
    )

    # =========================================================
    # PERMISSÕES POR PERFIL E BASE
    # =========================================================
    perfil = getattr(request.user, "perfil", None)

    if not request.user.is_superuser:
        if perfil is None:
            # Usuário sem perfil não pode visualizar checklists.
            checklists = checklists.none()

        elif perfil.role == "admin":
            # Admin visualiza somente os registros da própria empresa.
            if perfil.empresa_id:
                checklists = checklists.filter(
                    inventario__base__empresa_id=perfil.empresa_id
                )
            else:
                checklists = checklists.none()

        else:
            # Gestores e operadores visualizam somente as bases
            # associadas ao campo regionais do perfil.
            bases_permitidas = perfil.regionais.values_list(
                "id",
                flat=True,
            )

            checklists = checklists.filter(
                inventario__base_id__in=bases_permitidas
            )

    # =========================================================
    # PARÂMETROS DOS FILTROS
    # =========================================================
    pesquisa = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    data_inicio = request.GET.get("data_inicio", "").strip()
    data_fim = request.GET.get("data_fim", "").strip()
    por_pagina = request.GET.get("por_pagina", "10").strip()

    # =========================================================
    # PESQUISA GERAL
    # =========================================================
    if pesquisa:
        filtros_pesquisa = (
            Q(inventario__cliente__sigla__icontains=pesquisa)
            | Q(inventario__cliente__nome__icontains=pesquisa)
            | Q(inventario__base__nome__icontains=pesquisa)
            | Q(inventario__loja__icontains=pesquisa)
            | Q(criado_por__username__icontains=pesquisa)
        )

        if pesquisa.isdigit():
            filtros_pesquisa |= Q(id=int(pesquisa))

        checklists = checklists.filter(filtros_pesquisa)

    # =========================================================
    # FILTRO POR DATA INICIAL
    # =========================================================
    if data_inicio:
        try:
            data_inicio_convertida = datetime.strptime(
                data_inicio,
                "%Y-%m-%d",
            ).date()

            checklists = checklists.filter(
                data_inicio__date__gte=data_inicio_convertida
            )
        except ValueError:
            data_inicio = ""

    # =========================================================
    # FILTRO POR DATA FINAL
    # =========================================================
    if data_fim:
        try:
            data_fim_convertida = datetime.strptime(
                data_fim,
                "%Y-%m-%d",
            ).date()

            checklists = checklists.filter(
                data_inicio__date__lte=data_fim_convertida
            )
        except ValueError:
            data_fim = ""

    # =========================================================
    # CARDS DE RESUMO
    # Os contadores respeitam as permissões, pesquisa e período.
    # =========================================================
    queryset_resumo = checklists

    resumo = {
        "total": queryset_resumo.count(),
        "abertos": queryset_resumo.filter(
            status="ABERTO"
        ).count(),
        "em_execucao": queryset_resumo.filter(
            status="EM_EXECUCAO"
        ).count(),
        "finalizados": queryset_resumo.filter(
            status="FINALIZADO"
        ).count(),
    }

    # =========================================================
    # FILTRO POR STATUS
    # Aplicado após os cards para não zerar os outros contadores.
    # =========================================================
    status_validos = {
        "ABERTO",
        "EM_EXECUCAO",
        "FINALIZADO",
    }

    if status in status_validos:
        checklists = checklists.filter(status=status)
    else:
        status = ""

    # =========================================================
    # PAGINAÇÃO
    # =========================================================
    opcoes_por_pagina = {10, 25, 50}

    try:
        por_pagina_int = int(por_pagina)
    except (TypeError, ValueError):
        por_pagina_int = 10

    if por_pagina_int not in opcoes_por_pagina:
        por_pagina_int = 10

    paginator = Paginator(checklists, por_pagina_int)
    pagina = request.GET.get("page")
    page_obj = paginator.get_page(pagina)

    # Preserva os filtros durante a paginação.
    parametros = request.GET.copy()
    parametros.pop("page", None)

    context = {
        "checklists": page_obj,
        "page_obj": page_obj,
        "resumo": resumo,
        "pesquisa": pesquisa,
        "status_selecionado": status,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "por_pagina": por_pagina_int,
        "query_string": parametros.urlencode(),
    }

    return render(
        request,
        "insumos/lista_checklists.html",
        context,
    )

@login_required
def finalizar_checklist(request, pk):
    checklist = get_object_or_404(ChecklistDiario.objects.select_related('inventario__base', 'inventario__cliente'), pk=pk)

    # Verifica permissão (admin, gestor, ou responsável)
    perfil = request.user.perfil
    if not perfil.is_admin:
        if (
            checklist.inventario.base_id not in perfil.regionais_ids or
            checklist.inventario.base.empresa_id != perfil.empresa_id
        ):
            messages.error(request, 'Voce nao tem acesso a este checklist.')
            return redirect('insumos:lista_checklists')
        if not perfil.is_gestor and checklist.responsavel != request.user:
            messages.error(request, 'Você não tem permissão para finalizar este checklist.')
            return redirect('insumos:lista_checklists')

    # Verifica se já está finalizado
    if checklist.status == 'FINALIZADO':
        messages.warning(request, 'Este checklist já foi finalizado.')
        return redirect('insumos:lista_checklists')

    try:
        with transaction.atomic():
            ChecklistService.finalizar(checklist=checklist, usuario=request.user)
        messages.success(request, f'Checklist #{checklist.id} finalizado com sucesso!')
    except ValueError as e:
        messages.error(request, f'Erro ao finalizar checklist: {str(e)}')
    except Exception as e:
        messages.error(request, f'Erro inesperado: {str(e)}')

    return redirect('insumos:lista_checklists')

@login_required
def checklist_detail(request, pk):
    checklist = get_object_or_404(
        ChecklistDiario.objects.select_related(
            'inventario__cliente',
            'inventario__base',
            'inventario__base__empresa',
        ),
        pk=pk
    )

    perfil = request.user.perfil

    if not perfil.is_admin:
        if (
            checklist.inventario.base_id not in perfil.regionais_ids or
            checklist.inventario.base.empresa_id != perfil.empresa_id
        ):
            messages.error(request, 'Você não tem acesso a este checklist.')
            return redirect('insumos:lista_checklists')

    equipamentos_checklist = list(
        checklist.equipamentos_utilizados
        .select_related('equipamento__produto')
        .order_by('equipamento__produto__categoria', 'equipamento__patrimonio')
    )

    equipamentos_por_categoria = defaultdict(list)

    for item_equip in equipamentos_checklist:
        categoria = (
            item_equip.equipamento.produto.categoria
            if item_equip.equipamento.produto
            else 'Equipamentos'
        )
        chave = categoria.lower().replace(' ', '_')
        equipamentos_por_categoria[chave].append(item_equip)

    if request.method == 'POST':
        if checklist.status == 'FINALIZADO':
            messages.warning(request, 'Este checklist já foi finalizado.')
            return redirect('insumos:checklist_detail', pk=checklist.pk)

        try:
            with transaction.atomic():

                # INSUMOS
                for item in checklist.itens.select_related('insumo'):
                    retornada = request.POST.get(f'retornada_{item.id}', '').strip()

                    if retornada != '':
                        ChecklistService.atualizar_retorno_item(
                            item=item,
                            retornada=retornada,
                        )

                # EQUIPAMENTOS
                for chave, itens_categoria in equipamentos_por_categoria.items():
                    valor_retornado = request.POST.get(
                        f'equip_qtd_retornada_{chave}',
                        ''
                    ).strip()

                    if valor_retornado == '':
                        continue

                    try:
                        quantidade_retornada = int(valor_retornado)
                    except ValueError:
                        raise ValueError(
                            'Quantidade retornada de equipamentos inválida.'
                        )

                    quantidade_enviada = len(itens_categoria)

                    if quantidade_retornada < 0 or quantidade_retornada > quantidade_enviada:
                        raise ValueError(
                            'Quantidade retornada de equipamentos deve ficar entre zero e a quantidade enviada.'
                        )

                    ids_ocorrencia = {
                        int(equip_id)
                        for equip_id in request.POST.getlist(f'equip_ocorrencia_{chave}')
                        if str(equip_id).isdigit()
                    }

                    quantidade_divergente = quantidade_enviada - quantidade_retornada

                    if len(ids_ocorrencia) != quantidade_divergente:
                        produto = itens_categoria[0].equipamento.produto
                        categoria_label = (
                            produto.get_categoria_display()
                            if produto
                            else 'Equipamentos'
                        )

                        raise ValueError(
                            f'Informe exatamente {quantidade_divergente} equipamento(s) com ocorrência em {categoria_label}.'
                        )

                    for item_equip in itens_categoria:
                        if item_equip.id in ids_ocorrencia:
                            status_retorno = request.POST.get(
                                f'equip_ocorrencia_status_{item_equip.id}',
                                ''
                            ).strip()

                            observacao = request.POST.get(
                                f'equip_ocorrencia_obs_{item_equip.id}',
                                ''
                            ).strip()

                            if status_retorno in ('', 'PENDENTE', 'RETORNADO'):
                                raise ValueError(
                                    'Selecione o motivo da ocorrência do equipamento.'
                                )

                            if not observacao:
                                raise ValueError(
                                    'Informe a observação da ocorrência do equipamento.'
                                )

                            ChecklistService.resolver_retorno_equipamento(
                                item_equip=item_equip,
                                status_retorno=status_retorno,
                                observacao=observacao,
                                usuario=request.user,
                            )
                        else:
                            ChecklistService.resolver_retorno_equipamento(
                                item_equip=item_equip,
                                status_retorno='RETORNADO',
                                observacao='',
                                usuario=request.user,
                            )

                # TAGS
                for item_lote in checklist.lotes_tags_movimentados.select_related('lote'):
                    valor_final = request.POST.get(
                        f'tag_final_item_{item_lote.id}',
                        ''
                    ).strip()

                    if valor_final:
                        ChecklistService.atualizar_retorno_lote_tag(
                            item_lote=item_lote,
                            numero_final_utilizado=valor_final,
                        )

                if request.POST.get('acao') == 'finalizar':
                    ChecklistService.finalizar(
                        checklist=checklist,
                        usuario=request.user,
                    )
                    messages.success(
                        request,
                        f'Checklist #{checklist.id} finalizado com sucesso!'
                    )
                    return redirect('insumos:lista_checklists')

                messages.success(request, 'Retorno do checklist salvo com sucesso.')

        except ValueError as e:
            messages.error(request, f'Erro no retorno do checklist: {str(e)}')
        except Exception as e:
            messages.error(request, f'Erro inesperado: {str(e)}')

        return redirect('insumos:checklist_detail', pk=checklist.pk)

    tags = (
        checklist.lotes_tags_movimentados
        .select_related('lote', 'rolo')
        .order_by('lote__numero_inicial', 'numero_inicial_utilizado')
    )
    tags_insumos = (
        checklist.itens
        .select_related('insumo', 'insumo__categoria')
        .filter(insumo__categoria__nome='TAGS')
        .order_by('insumo__descricao')
    )

    equipamentos_grupos = []

    for chave, itens_categoria in equipamentos_por_categoria.items():
        resolvidos = [
            item
            for item in itens_categoria
            if item.status_retorno != 'PENDENTE'
        ]

        retornados = sum(
            1
            for item in itens_categoria
            if item.status_retorno == 'RETORNADO'
        )

        equipamentos_grupos.append({
            'key': chave,
            'categoria': (
                itens_categoria[0].equipamento.produto.get_categoria_display()
                if itens_categoria[0].equipamento.produto
                else 'Equipamentos'
            ),
            'enviados': len(itens_categoria),
            'retornados': retornados if resolvidos else len(itens_categoria),
            'itens': itens_categoria,
        })

    context = {
        'checklist': checklist,
        'itens': (
            checklist.itens
            .select_related('insumo', 'insumo__categoria')
            .exclude(insumo__categoria__nome='TAGS')
        ),
        'equipamentos': equipamentos_checklist,
        'equipamentos_grupos': equipamentos_grupos,
        'tags': tags,
        'tags_insumos': tags_insumos,
        'total_tags_utilizadas': sum(
            tag.quantidade_utilizada or 0
            for tag in tags
        ) + sum(
            item.quantidade_utilizada or 0
            for item in tags_insumos
        ),
    }

    return render(request, 'insumos/checklist_detail.html', context)

@login_required
def imprimir_checklist(request, pk):
    checklist = get_object_or_404(
        ChecklistDiario.objects.select_related(
            'inventario__cliente',
            'inventario__base',
            'inventario__base__empresa',
            'responsavel',
        ),
        pk=pk,
    )
    perfil = request.user.perfil
    if (
        not perfil.is_admin and
        (
            checklist.inventario.base_id not in perfil.bases_checklist_ids or
            checklist.inventario.base.empresa_id != perfil.empresa_id
        )
    ):
        messages.error(request, 'Você não tem acesso a este checklist.')
        return redirect('insumos:lista_checklists')

    itens = list(
        checklist.itens
        .select_related('insumo__categoria')
        .filter(quantidade_enviada__gt=0)
        .order_by('insumo__categoria__nome', 'insumo__descricao')
    )
    equipamentos = list(
        checklist.equipamentos_utilizados
        .select_related('equipamento__produto')
        .order_by(
            'equipamento__produto__categoria',
            'equipamento__produto__descricao',
            'equipamento__patrimonio',
        )
    )
    equipamentos_por_categoria = defaultdict(list)
    for item in equipamentos:
        categoria = (
            item.equipamento.produto.categoria
            if item.equipamento.produto_id
            else 'Equipamentos'
        )
        equipamentos_por_categoria[categoria].append(item)

    grupos_declaracao = defaultdict(float)
    for item in itens:
        categoria = (
            item.insumo.categoria.nome
            if item.insumo.categoria_id
            else 'Insumos'
        )
        grupos_declaracao[categoria] += float(item.quantidade_enviada)
    for categoria, lista in equipamentos_por_categoria.items():
        grupos_declaracao[categoria] += len(lista)

    linhas_por_categoria = defaultdict(list)
    for item in itens:
        categoria = (
            item.insumo.categoria.nome
            if item.insumo.categoria_id
            else 'INSUMOS'
        )
        linhas_por_categoria[categoria].append({
            'descricao': item.insumo.descricao,
            'quantidade_enviada': item.quantidade_enviada,
            'unidade': item.insumo.unidade_medida,
            'quantidade_retornada': (
                item.quantidade_retornada
                if item.status_retorno == 'CONFERIDO'
                else None
            ),
            'retorno_informado': item.status_retorno == 'CONFERIDO',
        })

    equipamentos_agrupados = defaultdict(lambda: {
        'quantidade_enviada': 0,
        'quantidade_retornada': 0,
    })
    for item in equipamentos:
        produto = item.equipamento.produto
        categoria = (
            produto.get_categoria_display()
            if produto
            else 'EQUIPAMENTOS'
        )
        descricao = produto.descricao if produto else 'Equipamento'
        chave = (categoria, descricao)
        equipamentos_agrupados[chave]['quantidade_enviada'] += 1
        if item.status_retorno == 'RETORNADO':
            equipamentos_agrupados[chave]['quantidade_retornada'] += 1

    for (categoria, descricao), quantidades in equipamentos_agrupados.items():
        linhas_por_categoria[categoria].append({
            'descricao': descricao,
            'quantidade_enviada': quantidades['quantidade_enviada'],
            'unidade': '',
            'quantidade_retornada': (
                quantidades['quantidade_retornada']
                if quantidades['quantidade_retornada']
                else None
            ),
            'retorno_informado': bool(quantidades['quantidade_retornada']),
        })

    grupos_checklist = [
        {'categoria': categoria, 'linhas': linhas}
        for categoria, linhas in linhas_por_categoria.items()
    ]

    if checklist.declaracao_quantidades:
        rotulos_declaracao = [
            ('departamento_pessoal', 'DEPARTAMENTO PESSOAL'),
            ('fios_cabos', 'FIOS E CABOS'),
            ('coletor_dados', 'COLETOR DE DADOS'),
            ('impressora', 'IMPRESSORA'),
            ('escada', 'ESCADA'),
            ('balanca', 'BALANÇA'),
            ('extensor_rede_carrinho', 'EXTENSOR DE REDE / CARRINHO'),
        ]
        grupos_declaracao = [
            (rotulo, checklist.declaracao_quantidades.get(chave, 0))
            for chave, rotulo in rotulos_declaracao
        ]
    else:
        grupos_declaracao = sorted(grupos_declaracao.items())

    dados_declaracao = {
        'cliente': checklist.inventario.cliente.sigla,
        'loja': str(checklist.inventario.loja),
        'data': checklist.inventario.data_inicio.strftime('%d/%m/%Y'),
        'endereco': checklist.inventario.endereco or '',
        'bairro': checklist.inventario.bairro or '',
        'cidade': checklist.inventario.cidade or '',
        'horario_entrega': (
            checklist.inventario.horario_ponto.strftime('%H:%M')
            if checklist.inventario.horario_ponto else ''
        ),
        'horario_inicio': (
            checklist.inventario.horario_inicio.strftime('%H:%M')
            if checklist.inventario.horario_inicio else ''
        ),
        'ponto_encontro': checklist.inventario.ponto_encontro or '',
        'transporte': checklist.transporte or '',
    }
    dados_declaracao.update(checklist.declaracao_dados or {})

    return render(
        request,
        'insumos/checklist_impressao.html',
        {
            'checklist': checklist,
            'itens': itens,
            'equipamentos': equipamentos,
            'equipamentos_por_categoria': dict(equipamentos_por_categoria),
            'grupos_declaracao': grupos_declaracao,
            'dados_declaracao': dados_declaracao,
            'grupos_checklist': grupos_checklist,
        },
    )

@login_required
def exportar_checklist_modelo(request, pk):
    checklist = get_object_or_404(
        ChecklistDiario.objects.select_related(
            'inventario__cliente',
            'inventario__base',
            'inventario__base__empresa',
        ),
        pk=pk,
    )
    perfil = request.user.perfil
    if (
        not perfil.is_admin and
        (
            checklist.inventario.base_id not in perfil.bases_checklist_ids or
            checklist.inventario.base.empresa_id != perfil.empresa_id
        )
    ):
        messages.error(request, 'Você não tem acesso a este checklist.')
        return redirect('insumos:lista_checklists')

    modelo = (
        Path(settings.BASE_DIR) /
        'insumos' /
        'templates_xlsx' /
        'checklist_declaracao_modelo.xlsx'
    )
    workbook = openpyxl.load_workbook(modelo)
    planilha_checklist = workbook['Check - List']
    planilha_declaracao = workbook['Declaração']

    for planilha in list(workbook.worksheets):
        if planilha not in (planilha_checklist, planilha_declaracao):
            workbook.remove(planilha)
    workbook._sheets = [planilha_declaracao, planilha_checklist]
    workbook.active = 0

    inventario = checklist.inventario
    data_formatada = inventario.data_inicio.strftime('%d/%m/%Y')
    cliente = inventario.cliente.sigla
    loja = f'Loja {inventario.loja}'

    planilha_checklist['C3'] = cliente
    planilha_checklist['E3'] = loja
    planilha_checklist['G3'] = data_formatada
    planilha_checklist['C5'] = inventario.endereco or ''
    planilha_checklist['C7'] = inventario.bairro or ''
    planilha_checklist['G7'] = inventario.cidade or ''
    planilha_checklist['E76'] = checklist.quantidade_volumes

    def chave_descricao(valor):
        texto = unicodedata.normalize('NFKD', str(valor or ''))
        return ''.join(
            caractere for caractere in texto
            if not unicodedata.combining(caractere)
        ).strip().casefold()

    quantidades_por_descricao = defaultdict(Decimal)
    for item in checklist.itens.select_related('insumo'):
        quantidades_por_descricao[
            chave_descricao(item.insumo.descricao)
        ] += item.quantidade_enviada

    equipamentos_por_categoria = defaultdict(int)
    for item in checklist.equipamentos_utilizados.select_related(
        'equipamento__produto'
    ):
        categoria = (
            item.equipamento.produto.categoria
            if item.equipamento.produto_id
            else 'Equipamentos'
        )
        equipamentos_por_categoria[chave_descricao(categoria)] += 1

    linha_equipamento = {
        'routers': 38,
        'roteadores': 38,
        'coletores': 51,
        'notebooks': 61,
        'impressoras': 63,
    }
    for linha in range(11, 75):
        descricao = planilha_checklist.cell(linha, 3).value
        quantidade = quantidades_por_descricao.get(chave_descricao(descricao), 0)
        planilha_checklist.cell(linha, 5).value = quantidade or ''
    for categoria, linha in linha_equipamento.items():
        quantidade = equipamentos_por_categoria.get(categoria, 0)
        if quantidade:
            planilha_checklist.cell(linha, 5).value = quantidade

    dados_declaracao = {
        'cliente': cliente,
        'loja': loja,
        'data': data_formatada,
        'endereco': inventario.endereco or '',
        'bairro': inventario.bairro or '',
        'cidade': inventario.cidade or '',
        'horario_entrega': (
            inventario.horario_ponto.strftime('%H:%M')
            if inventario.horario_ponto else ''
        ),
        'horario_inicio': (
            inventario.horario_inicio.strftime('%H:%M')
            if inventario.horario_inicio else ''
        ),
        'ponto_encontro': inventario.ponto_encontro or '',
        'transporte': checklist.transporte or '',
    }
    dados_declaracao.update(checklist.declaracao_dados or {})
    planilha_declaracao['C4'] = dados_declaracao['cliente']
    planilha_declaracao['E4'] = dados_declaracao['loja']
    planilha_declaracao['G4'] = dados_declaracao['data']
    planilha_declaracao['C6'] = dados_declaracao['endereco']
    planilha_declaracao['C8'] = dados_declaracao['bairro']
    planilha_declaracao['F8'] = dados_declaracao['cidade']
    planilha_declaracao['C10'] = dados_declaracao['horario_entrega']
    planilha_declaracao['G10'] = dados_declaracao['horario_inicio']
    planilha_declaracao['D12'] = dados_declaracao['ponto_encontro']
    planilha_declaracao['D14'] = dados_declaracao['transporte']

    grupos_declaracao = {
        27: Decimal('0'),
        29: Decimal('0'),
        31: Decimal(equipamentos_por_categoria.get('coletores', 0)),
        33: Decimal(equipamentos_por_categoria.get('impressoras', 0)),
        35: quantidades_por_descricao.get(chave_descricao('Escada'), Decimal('0')),
        37: quantidades_por_descricao.get(chave_descricao('Balança'), Decimal('0')),
        39: quantidades_por_descricao.get(
            chave_descricao('Extensor de Rede / Carrinho'),
            Decimal('0'),
        ),
    }
    for item in checklist.itens.select_related('insumo__categoria'):
        categoria = chave_descricao(
            item.insumo.categoria.nome
            if item.insumo.categoria_id else ''
        )
        if categoria == chave_descricao('Departamento Pessoal'):
            grupos_declaracao[27] += item.quantidade_enviada
        elif categoria == chave_descricao('Fios e Cabos'):
            grupos_declaracao[29] += item.quantidade_enviada

    if checklist.declaracao_quantidades:
        chaves_por_linha = {
            27: 'departamento_pessoal',
            29: 'fios_cabos',
            31: 'coletor_dados',
            33: 'impressora',
            35: 'escada',
            37: 'balanca',
            39: 'extensor_rede_carrinho',
        }
        grupos_declaracao = {
            linha: Decimal(str(
                checklist.declaracao_quantidades.get(chave, 0)
            ))
            for linha, chave in chaves_por_linha.items()
        }

    for linha, quantidade in grupos_declaracao.items():
        planilha_declaracao.cell(linha, 6).value = quantidade or ''
    planilha_declaracao['F41'] = checklist.quantidade_volumes

    arquivo = BytesIO()
    workbook.save(arquivo)
    arquivo.seek(0)
    response = HttpResponse(
        arquivo.getvalue(),
        content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ),
    )
    response['Content-Disposition'] = (
        f'attachment; filename="checklist_{checklist.pk}_modelo_oficial.xlsx"'
    )
    return response

@login_required
def editar_itens_checklist(request, pk):
    checklist = get_object_or_404(ChecklistDiario.objects.select_related('inventario__base'), pk=pk)

    # Verifica permissão
    perfil = request.user.perfil
    if (
        not perfil.is_admin and
        (
            checklist.inventario.base_id not in perfil.regionais_ids or
            checklist.inventario.base.empresa_id != perfil.empresa_id
        )
    ):
        messages.error(request, 'Você não tem acesso a este checklist.')
        return redirect('insumos:lista_checklists')

    if checklist.status == 'FINALIZADO':
        messages.warning(request, 'Este checklist já está finalizado.')
        return redirect('insumos:lista_checklists')

    if request.method == 'POST':
        for item in checklist.itens.select_related('insumo'):
            retornada = request.POST.get(f'retornada_{item.id}', '').strip()

            try:
                if retornada == '':
                    continue

                ChecklistService.atualizar_retorno_item(
                    item=item,
                    retornada=retornada,
                )
            except ValueError as e:
                messages.error(request, f'Erro no item "{item.insumo.descricao}": {str(e)}')
                return redirect('insumos:editar_itens_checklist', pk=checklist.id)

        messages.success(request, 'Itens do checklist atualizados com sucesso!')
        return redirect('insumos:checklist_detail', pk=checklist.id)

    context = {
        'checklist': checklist,
        'itens': checklist.itens.select_related('insumo'),
    }
    return render(request, 'insumos/editar_itens_checklist.html', context)

@login_required
def ultimo_checklist_por_loja(request):
    inventario_id = request.GET.get('inventario')
    if not inventario_id:
        return JsonResponse({'error': 'Inventário não informado'}, status=400)

    inventario = get_object_or_404(
        Inventario.objects.select_related('cliente', 'base', 'base__empresa'),
        pk=inventario_id,
    )
    perfil = request.user.perfil
    if not perfil.is_admin and (
        inventario.base_id not in perfil.bases_checklist_ids or
        inventario.base.empresa_id != perfil.empresa_id
    ):
        return JsonResponse({'error': 'Base não permitida para este usuário'}, status=403)

    anteriores = ChecklistDiario.objects.filter(
        inventario__cliente_id=inventario.cliente_id,
        inventario__loja__iexact=inventario.loja,
        inventario__base_id=inventario.base_id,
        inventario__data_inicio__lt=inventario.data_inicio,
        status__in=['FINALIZADO', 'EM_EXECUCAO'],
    ).exclude(inventario_id=inventario.id).select_related('inventario')

    # Usa o preenchimento imediatamente anterior, mesmo que o retorno ainda esteja em execução.
    ultimo = anteriores.order_by('-inventario__data_inicio', '-data_inicio').first()

    if not ultimo:
        return JsonResponse({'dados': None})

    # Monta resposta
    data = {
        'checklist_id': ultimo.id,
        'inventario_id': ultimo.inventario_id,
        'data': ultimo.data_inicio.strftime('%d/%m/%Y %H:%M'),
        'status': ultimo.status,
        'insumos': [],
        'equipamentos': [],
        'tags': []
    }

    # Insumos
    for item in ultimo.itens.select_related('insumo'):
        data['insumos'].append({
            'insumo_id': item.insumo.id,
            'enviada': float(item.quantidade_enviada),
            'utilizada': float(item.quantidade_utilizada or 0),
            'retornada': float(item.quantidade_retornada or 0),
            'perdida': float(item.quantidade_perdida or 0),
        })

    # Equipamentos
    categorias = {
        'coletores': 'coletor',
        'coletor': 'coletor',
        'impressoras': 'impressora',
        'impressora': 'impressora',
        'notebooks': 'notebook',
        'notebook': 'notebook',
        'routers': 'router',
        'router': 'router',
        'roteadores': 'router',
        'roteador': 'router',
    }
    for eq in ultimo.equipamentos_utilizados.select_related('equipamento__produto'):
        equipamento = eq.equipamento
        if equipamento.status != 'ATIVO' or equipamento.regional_id != inventario.base_id:
            continue
        categoria = categorias.get((equipamento.produto.categoria or '').strip().lower())
        if not categoria:
            continue
        data['equipamentos'].append({
            'id': equipamento.id,
            'categoria': categoria,
            'tag_saida': eq.tag_saida,
        })

    # Tags (LoteTag) - se houver
    for tag in ultimo.lotes_tags_movimentados.select_related('lote', 'rolo'):
        if not tag.rolo or tag.rolo.status not in ['DISPONIVEL', 'EM_USO']:
            continue
        if tag.lote.base_id != inventario.base_id:
            continue
        data['tags'].append({
            'lote_id': tag.lote.id,
            'rolo_id': tag.rolo_id,
            'inicial_sugerido': tag.rolo.numero_atual,
        })

    return JsonResponse({'dados': data})

@login_required
def editar_checklist(request, pk):
    checklist = get_object_or_404(ChecklistDiario.objects.select_related(
        'inventario__base', 'inventario__cliente'
    ), pk=pk)
    perfil = request.user.perfil

    # Verifica permissão (admin, gestor, ou responsável)
    if not perfil.is_admin:
        if not perfil.is_gestor and checklist.responsavel != request.user:
            messages.error(request, 'Você não tem permissão para editar este checklist.')
            return redirect('insumos:lista_checklists')

    if checklist.status == 'FINALIZADO':
        messages.warning(request, 'Checklist já finalizado, não pode ser editado.')
        return redirect('insumos:checklist_detail', pk=checklist.pk)

    if request.method == 'POST':
        try:
            with transaction.atomic():
                campos_declaracao = {
                    'departamento_pessoal': 'declaracao_departamento_pessoal',
                    'fios_cabos': 'declaracao_fios_cabos',
                    'coletor_dados': 'declaracao_coletor_dados',
                    'impressora': 'declaracao_impressora',
                    'escada': 'declaracao_escada',
                    'balanca': 'declaracao_balanca',
                    'extensor_rede_carrinho': 'declaracao_extensor_rede_carrinho',
                }
                declaracao_enviada = any(
                    nome_post in request.POST
                    for nome_post in campos_declaracao.values()
                )
                if declaracao_enviada:
                    declaracao_quantidades = {}
                    for chave, nome_post in campos_declaracao.items():
                        try:
                            quantidade = int(
                                request.POST.get(nome_post, '0') or '0'
                            )
                        except (TypeError, ValueError) as exc:
                            raise ValueError(
                                'As quantidades da declaração devem ser números inteiros.'
                            ) from exc
                        if quantidade < 0:
                            raise ValueError(
                                'As quantidades da declaração não podem ser negativas.'
                            )
                        declaracao_quantidades[chave] = quantidade
                else:
                    declaracao_quantidades = checklist.declaracao_quantidades
                try:
                    quantidade_volumes = int(
                        request.POST.get(
                            'quantidade_volumes',
                            checklist.quantidade_volumes,
                        )
                    )
                except (TypeError, ValueError) as exc:
                    raise ValueError(
                        'Informe uma quantidade de volumes válida.'
                    ) from exc
                if quantidade_volumes <= 0 or quantidade_volumes > 9999:
                    raise ValueError(
                        'Informe uma quantidade de volumes entre 1 e 9999.'
                    )
                checklist.quantidade_volumes = quantidade_volumes
                checklist.declaracao_quantidades = declaracao_quantidades
                checklist.transporte = request.POST.get(
                    'transporte',
                    checklist.transporte,
                ).strip()
                checklist.observacao = request.POST.get(
                    'observacao',
                    checklist.observacao,
                ).strip()
                inventario = checklist.inventario
                dados_atuais = checklist.declaracao_dados or {}

                def valor_declaracao(nome_post, chave, padrao=''):
                    return request.POST.get(
                        nome_post,
                        dados_atuais.get(chave, padrao),
                    ).strip()

                checklist.declaracao_dados = {
                    'cliente': valor_declaracao(
                        'declaracao_cliente', 'cliente', inventario.cliente.sigla
                    ),
                    'loja': valor_declaracao(
                        'declaracao_loja', 'loja', str(inventario.loja)
                    ),
                    'data': valor_declaracao(
                        'declaracao_data', 'data',
                        inventario.data_inicio.strftime('%d/%m/%Y')
                    ),
                    'endereco': valor_declaracao(
                        'declaracao_endereco', 'endereco', inventario.endereco or ''
                    ),
                    'bairro': valor_declaracao(
                        'declaracao_bairro', 'bairro', inventario.bairro or ''
                    ),
                    'cidade': valor_declaracao(
                        'declaracao_cidade', 'cidade', inventario.cidade or ''
                    ),
                    'horario_entrega': valor_declaracao(
                        'horario_ponto', 'horario_entrega',
                        inventario.horario_ponto.strftime('%H:%M')
                        if inventario.horario_ponto else ''
                    ),
                    'horario_inicio': valor_declaracao(
                        'horario_inicio', 'horario_inicio',
                        inventario.horario_inicio.strftime('%H:%M')
                        if inventario.horario_inicio else ''
                    ),
                    'ponto_encontro': valor_declaracao(
                        'ponto_encontro', 'ponto_encontro',
                        inventario.ponto_encontro or ''
                    ),
                    'transporte': checklist.transporte,
                }
                checklist.save(update_fields=[
                    'quantidade_volumes',
                    'declaracao_quantidades',
                    'declaracao_dados',
                    'transporte',
                    'observacao',
                ])

                def ler_horario(nome):
                    if nome not in request.POST:
                        return getattr(inventario, nome)
                    valor = request.POST.get(nome, '').strip()
                    if not valor:
                        return None
                    try:
                        return datetime.strptime(valor, '%H:%M').time()
                    except ValueError as exc:
                        raise ValueError(
                            f'Informe um horário válido para {nome.replace("_", " ")}.'
                        ) from exc

                inventario.ponto_encontro = request.POST.get(
                    'ponto_encontro', inventario.ponto_encontro or ''
                ).strip()
                inventario.horario_ponto = ler_horario('horario_ponto')
                inventario.horario_inicio = ler_horario('horario_inicio')
                inventario.save(update_fields=[
                    'ponto_encontro',
                    'horario_ponto',
                    'horario_inicio',
                ])

                # Atualizar insumos (quantidades utilizada, retornada, perdida)
                for item in checklist.itens.select_related('insumo'):
                    utilizada = request.POST.get(f'insumo_{item.insumo.id}_utilizada', 0)
                    retornada = request.POST.get(f'insumo_{item.insumo.id}_retornada', 0)
                    perdida = request.POST.get(f'insumo_{item.insumo.id}_perdida', 0)
                    # Atualiza o item
                    item.quantidade_utilizada = Decimal(utilizada or 0)
                    item.quantidade_retornada = Decimal(retornada or 0)
                    item.quantidade_perdida = Decimal(perdida or 0)
                    item.save(update_fields=['quantidade_utilizada', 'quantidade_retornada', 'quantidade_perdida'])

                # Atualizar equipamentos (tag_volta = retorno confirmado)
                for eq in checklist.equipamentos_utilizados.all():
                    # Identifica a categoria para buscar o campo de retorno
                    categoria = eq.equipamento.produto.categoria.lower()
                    retorno = request.POST.get(f'retorno_equip_{categoria}', '')
                    if retorno:
                        eq.tag_volta = retorno
                        eq.save(update_fields=['tag_volta'])

                messages.success(request, 'Checklist atualizado com sucesso!')
                return redirect('insumos:checklist_detail', pk=checklist.pk)

        except Exception as e:
            messages.error(request, f'Erro ao atualizar: {str(e)}')


    # Prepara listas de equipamentos (como na view de criação)
    from estoque.models import Equipamento
    from insumos.models import Insumo

    regionais_ids = perfil.bases_checklist_ids
    if perfil.is_admin:
        equipamentos = Equipamento.objects.filter(
            status='ATIVO', finalidade=Equipamento.Finalidade.OPERACIONAL
        )
        lotes_tags = RoloTag.objects.filter(status__in=['DISPONIVEL', 'EM_USO'], lote__ativo=True).select_related('lote', 'lote__base')
    else:
        equipamentos = Equipamento.objects.filter(
            status='ATIVO',
            finalidade=Equipamento.Finalidade.OPERACIONAL,
            regional_id__in=regionais_ids,
            regional__empresa=perfil.empresa,
        )
        lotes_tags = RoloTag.objects.filter(
            status__in=['DISPONIVEL', 'EM_USO'],
            lote__ativo=True,
            lote__base_id__in=regionais_ids,
            lote__base__empresa=perfil.empresa,
        ).select_related('lote', 'lote__base')

    # Obtém itens do checklist para pré-preencher
    itens_checklist = []

    # Insumos já existentes no checklist
    insumos_do_checklist = checklist.itens.select_related('insumo')

    context = {
        'checklist': checklist,
        'coletores': equipamentos.filter(produto__categoria='Coletores'),
        'impressoras': equipamentos.filter(produto__categoria='Impressoras'),
        'notebooks': equipamentos.filter(produto__categoria='Notebooks'),
        'routers': equipamentos.filter(produto__categoria='Routers'),
        'inventarios': [checklist.inventario],  # apenas o inventário atual
        'insumos': insumos_do_checklist,  # insumos já adicionados
        'lotes_tags': lotes_tags,
        'lotes_tags_selecionados': checklist.lotes_tags_movimentados.select_related('lote'),
        'url_name': 'editar_checklist',
        'editando': True,  # flag para indicar que é edição
        'equipamentos_selecionados': checklist.equipamentos_utilizados.all(),
    }
    return render(request, 'estoque/checklist.html', context)

@login_required
@role_required('admin', 'gestor')
def ajustar_estoque_insumo(request):

    if request.method != 'POST':
        return redirect('insumos:estoque_insumos')

    base_id = request.POST.get('base_id')
    insumo_id = request.POST.get('insumo_id')
    saldo_real = request.POST.get('saldo_real')
    motivo = request.POST.get('motivo')
    senha = request.POST.get('senha')

    if not request.user.check_password(senha):
        messages.error(request, 'Senha inválida. Ajuste não realizado.')
        return redirect('insumos:estoque_insumos')

    if not motivo:
        messages.error(request, 'Informe o motivo do ajuste.')
        return redirect('insumos:estoque_insumos')

    try:
        saldo_real = Decimal(saldo_real)
    except:
        messages.error(request, 'Saldo informado inválido.')
        return redirect('insumos:estoque_insumos')

    base = get_object_or_404(Base, id=base_id)
    insumo = get_object_or_404(Insumo, id=insumo_id)
    saldo_anterior = MovimentacaoService.saldo(base, insumo)

    MovimentacaoService.ajuste(
        base=base,
        insumo=insumo,
        saldo_real=saldo_real,
        usuario=request.user,
        observacao=motivo,
    )

    admins = User.objects.filter(perfil__role='admin', is_active=True).distinct()
    if admins.exists():
        comunicado = Comunicado.objects.create(
            titulo='Ajuste de estoque de insumo',
            mensagem=(
                f'O estoque do insumo {insumo.descricao} foi ajustado.\n\n'
                f'Base: {base.nome}\n'
                f'Saldo anterior: {saldo_anterior}\n'
                f'Saldo ajustado: {saldo_real}\n'
                f'Motivo: {motivo}\n'
                f'Usuário: {request.user.get_username()}'
            ),
            tipo='OPERACIONAL',
            criado_por=request.user,
            enviar_para_todos=False,
            permitir_limpar=False,
        )
        comunicado.usuarios.set(admins)

    messages.success(
        request,
        f'Estoque de {insumo.descricao} ajustado com sucesso.'
    )

    return redirect('insumos:estoque_insumos')
