import uuid
from decimal import Decimal, InvalidOperation
from zipfile import BadZipFile

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.paginator import Paginator
from django.db.models import Count, DecimalField, ExpressionWrapper, F, Q, Sum, Value
from django.db import transaction
from django.db.models.functions import Coalesce
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.http import require_POST
from django.utils.translation import gettext as _
from django.utils.http import url_has_allowed_host_and_scheme

from auditorias.services.visibilidade_estoque_service import VisibilidadeEstoqueAuditoriaService
from compras.forms import AquisicaoForm, ImportacaoPrecificacaoForm, ItemAquisicaoForm, RemessaForm
from compras.models import Aquisicao, CodigoCatalogo, ItemRemessaCompra, RemessaCompra
from compras.policies import AquisicaoAccessPolicy
from compras.services import AquisicaoService, ProdutoPrecoService, RemessaCompraService
from estoque.forms import ProdutoForm
from estoque.models import Base, Equipamento, Produto
from estoque.policies.compras import ComprasAccessPolicy
from insumos.models import CategoriaInsumo, FornecedorInsumo, Insumo, SaldoInsumoBase
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException


@login_required
def lista_aquisicoes(request):
    return render(request, 'compras/aquisicao_lista.html', {
        'aquisicoes': AquisicaoAccessPolicy.queryset(request.user).select_related(
            'empresa', 'fornecedor', 'cadastrado_por'
        ).prefetch_related('itens')[:250],
        'pode_gerenciar': AquisicaoAccessPolicy.pode_gerenciar(request.user),
    })


@login_required
def criar_aquisicao(request):
    if not AquisicaoAccessPolicy.pode_gerenciar(request.user):
        raise PermissionDenied
    form = AquisicaoForm(request.POST or None, request.FILES or None)
    item_form = ItemAquisicaoForm(request.POST or None)
    form.fields['empresa'].queryset = ComprasAccessPolicy.empresas(request.user)
    if request.method == 'POST' and form.is_valid() and item_form.is_valid():
        dados = form.cleaned_data.copy()
        empresa = dados.pop('empresa')
        fornecedor = dados.pop('fornecedor')
        item = item_form.cleaned_data
        try:
            aquisicao = AquisicaoService.criar(
                empresa=empresa, fornecedor=fornecedor, usuario=request.user,
                itens=[item], **dados,
            )
        except (ValidationError, PermissionDenied) as exc:
            form.add_error(None, exc)
        else:
            messages.success(request, 'Aquisição cadastrada com rastreabilidade financeira.')
            return redirect('compras:aquisicao_detalhe', pk=aquisicao.pk)
    return render(request, 'compras/aquisicao_form.html', {'form': form, 'item_form': item_form})


@login_required
def detalhe_aquisicao(request, pk):
    aquisicao = get_object_or_404(
        AquisicaoAccessPolicy.queryset(request.user).select_related(
            'empresa', 'fornecedor', 'cadastrado_por', 'aprovado_por'
        ).prefetch_related('itens__produto', 'itens__insumo', 'eventos'),
        pk=pk,
    )
    return render(request, 'compras/aquisicao_detalhe.html', {
        'aquisicao': aquisicao,
        'pode_gerenciar': AquisicaoAccessPolicy.pode_gerenciar(request.user),
    })


@login_required
@require_POST
def aprovar_aquisicao(request, pk):
    aquisicao = get_object_or_404(AquisicaoAccessPolicy.queryset(request.user), pk=pk)
    try:
        AquisicaoService.aprovar(aquisicao, request.user)
        messages.success(request, 'Aquisição aprovada.')
    except (ValidationError, PermissionDenied) as exc:
        messages.error(request, str(exc))
    return redirect('compras:aquisicao_detalhe', pk=pk)


@login_required
def documento_aquisicao(request, pk, tipo):
    aquisicao = get_object_or_404(AquisicaoAccessPolicy.queryset(request.user), pk=pk)
    campo = {'danfe': aquisicao.arquivo_danfe_pdf, 'xml': aquisicao.arquivo_xml_nfe}.get(tipo)
    if not campo:
        raise PermissionDenied('Documento inexistente ou não autorizado.')
    return FileResponse(campo.open('rb'), as_attachment=True, filename=campo.name.rsplit('/', 1)[-1])


@login_required
def lista_remessas(request):
    remessas = AquisicaoAccessPolicy.remessas(request.user).select_related(
        'empresa', 'base_origem', 'base_destino', 'aquisicao'
    ).prefetch_related('itens')
    return render(request, 'compras/remessa_lista.html', {
        'remessas': remessas[:250],
        'pode_criar': ComprasAccessPolicy.pode_criar_remessa(request.user),
    })


@login_required
def criar_remessa(request):
    if not ComprasAccessPolicy.pode_criar_remessa(request.user):
        raise PermissionDenied
    form = RemessaForm(request.POST or None, user=request.user)
    if request.method == 'POST' and form.is_valid():
        dados = form.cleaned_data
        item = {
            'insumo': dados['insumo'], 'equipamento': dados['equipamento'],
            'item_aquisicao': dados['item_aquisicao'],
            'quantidade_prevista': dados['quantidade'],
            'custo_unitario_snapshot': dados['custo_unitario'],
        }
        try:
            remessa = RemessaCompraService.criar(
                empresa=dados['empresa'], fluxo=dados['fluxo'],
                aquisicao=dados['aquisicao'], base_origem=dados['base_origem'],
                base_destino=dados['base_destino'], usuario=request.user,
                previsao_chegada=dados['previsao_chegada'], observacao=dados['observacao'],
                itens=[item],
            )
        except (ValidationError, PermissionDenied, ValueError) as exc:
            form.add_error(None, exc)
        else:
            messages.success(request, f'Remessa {remessa.protocolo} criada.')
            return redirect('compras:remessa_detalhe', pk=remessa.pk)
    return render(request, 'compras/remessa_form.html', {'form': form})


@login_required
def detalhe_remessa(request, pk):
    remessa = get_object_or_404(
        AquisicaoAccessPolicy.remessas(request.user).select_related(
            'empresa', 'base_origem', 'base_destino', 'aquisicao', 'criada_por'
        ).prefetch_related('itens__insumo', 'itens__equipamento__produto', 'recebimentos__linhas'),
        pk=pk,
    )
    return render(request, 'compras/remessa_detalhe.html', {
        'remessa': remessa,
        'pode_enviar': ComprasAccessPolicy.pode_criar_remessa(request.user),
        'pode_confirmar': AquisicaoAccessPolicy.pode_confirmar(request.user, remessa),
        'idempotency_key': uuid.uuid4(),
    })


@login_required
@require_POST
def enviar_remessa(request, pk):
    remessa = get_object_or_404(AquisicaoAccessPolicy.remessas(request.user), pk=pk)
    try:
        RemessaCompraService.enviar(remessa, request.user, request.POST.get('codigo_rastreio', ''))
        messages.success(request, 'Remessa enviada para conferência do destino.')
    except (ValidationError, PermissionDenied) as exc:
        messages.error(request, str(exc))
    return redirect('compras:remessa_detalhe', pk=pk)


@login_required
@require_POST
def confirmar_remessa(request, pk):
    remessa = get_object_or_404(AquisicaoAccessPolicy.remessas(request.user), pk=pk)
    linhas = []
    for item in remessa.itens.all():
        linhas.append({
            'item_id': item.pk,
            'quantidade_recebida': request.POST.get(f'recebida_{item.pk}', 0),
            'quantidade_avariada': request.POST.get(f'avariada_{item.pk}', 0),
            'quantidade_faltante': request.POST.get(f'faltante_{item.pk}', 0),
            'observacao': request.POST.get(f'observacao_{item.pk}', ''),
        })
    try:
        RemessaCompraService.confirmar(
            remessa=remessa, usuario=request.user,
            idempotency_key=request.POST.get('idempotency_key'), linhas=linhas,
            finalizar=request.POST.get('finalizar') == '1',
            observacao=request.POST.get('observacao', ''),
        )
        messages.success(request, 'Conferência registrada sem duplicar entradas.')
    except (ValidationError, PermissionDenied, ValueError) as exc:
        messages.error(request, str(exc))
    return redirect('compras:remessa_detalhe', pk=pk)


@login_required
def valores_insumos(request):
    if not ComprasAccessPolicy.pode_visualizar_valores(request.user):
        raise PermissionDenied
    saldos = SaldoInsumoBase.objects.select_related('base__empresa', 'insumo__categoria')
    bases = ComprasAccessPolicy.bases(request.user)
    if not request.user.perfil.is_admin:
        saldos = saldos.filter(base__in=bases)
    base_id = request.GET.get('base', '').strip()
    categoria_id = request.GET.get('categoria', '').strip()
    insumo_id = request.GET.get('insumo', '').strip()
    busca = request.GET.get('q', '').strip()
    if base_id and base_id.isdigit():
        saldos = saldos.filter(base_id=base_id)
    if categoria_id.isdigit():
        saldos = saldos.filter(insumo__categoria_id=categoria_id)
    if insumo_id.isdigit():
        saldos = saldos.filter(insumo_id=insumo_id)
    if busca:
        saldos = saldos.filter(
            Q(insumo__descricao__icontains=busca)
            | Q(insumo__categoria__nome__icontains=busca)
            | Q(base__nome__icontains=busca)
        )
    valor = ExpressionWrapper(
        F('saldo') * F('custo_medio'),
        output_field=DecimalField(max_digits=28, decimal_places=6),
    )
    disponivel = ExpressionWrapper(
        F('saldo') - F('saldo_reservado'),
        output_field=DecimalField(max_digits=14, decimal_places=2),
    )
    saldos = saldos.annotate(valor_calculado=valor, saldo_disponivel_calculado=disponivel)
    resumo = saldos.aggregate(
        total=Coalesce(Sum('valor_calculado'), Value(Decimal('0'))),
        saldo=Coalesce(Sum('saldo'), Value(Decimal('0'))),
        reservado=Coalesce(Sum('saldo_reservado'), Value(Decimal('0'))),
        disponivel=Coalesce(Sum('saldo_disponivel_calculado'), Value(Decimal('0'))),
    )
    sem_preco = saldos.filter(saldo__gt=0, custo_medio=0).count()
    skus = saldos.values('insumo_id').distinct().count()
    bases_com_saldo = saldos.filter(saldo__gt=0).values('base_id').distinct().count()
    page_obj = Paginator(
        saldos.order_by('base__nome', 'insumo__descricao'), 20
    ).get_page(request.GET.get('page'))
    query_params = request.GET.copy()
    query_params.pop('page', None)
    insumos = Insumo.objects.filter(ativo=True)
    if categoria_id.isdigit():
        insumos = insumos.filter(categoria_id=categoria_id)
    return render(request, 'compras/valores_insumos.html', {
        'saldos': page_obj.object_list,
        'page_obj': page_obj,
        'query_string': query_params.urlencode(),
        'bases': bases,
        'categorias': CategoriaInsumo.objects.order_by('nome'),
        'insumos': insumos.select_related('categoria').order_by('categoria__nome', 'descricao'),
        'total': resumo['total'],
        'saldo_total': resumo['saldo'],
        'saldo_reservado': resumo['reservado'],
        'saldo_disponivel': resumo['disponivel'],
        'skus': skus,
        'bases_com_saldo': bases_com_saldo,
        'sem_preco': sem_preco,
        'filtros': {
            'base': base_id, 'categoria': categoria_id,
            'insumo': insumo_id, 'q': busca,
        },
        'grafico_bases': list(
            saldos.values('base__nome').annotate(total=Sum('valor_calculado')).order_by('-total')[:10]
        ),
        'grafico_categorias': list(
            saldos.values('insumo__categoria__nome').annotate(
                total=Sum('valor_calculado')
            ).order_by('-total')
        ),
        'grafico_insumos': list(
            saldos.values('insumo__descricao').annotate(
                total=Sum('valor_calculado')
            ).order_by('-total')[:10]
        ),
    })


@login_required
def valores_equipamentos(request):
    if not ComprasAccessPolicy.pode_visualizar_valores(request.user):
        raise PermissionDenied
    equipamentos = VisibilidadeEstoqueAuditoriaService.ocultar_equipamentos(
        Equipamento.objects.select_related('produto', 'regional__empresa', 'fornecedor')
    )
    bases = ComprasAccessPolicy.bases(request.user)
    if not request.user.perfil.is_admin:
        equipamentos = equipamentos.filter(regional__in=bases)
    base_id = request.GET.get('base')
    if base_id and base_id.isdigit():
        equipamentos = equipamentos.filter(regional_id=base_id)
    categoria = request.GET.get('categoria', '').strip()
    produto_id = request.GET.get('equipamento', '').strip()
    busca = request.GET.get('q', '').strip()
    if categoria:
        equipamentos = equipamentos.filter(produto__categoria=categoria)
    if produto_id.isdigit():
        equipamentos = equipamentos.filter(produto_id=produto_id)
    if busca:
        equipamentos = equipamentos.filter(
            Q(produto__descricao__icontains=busca)
            | Q(produto__modelo__icontains=busca)
            | Q(patrimonio__icontains=busca)
            | Q(numero_serie__icontains=busca)
            | Q(codigo__icontains=busca)
            | Q(regional__nome__icontains=busca)
        )
    valor = Coalesce(
        'custo_aquisicao', 'produto__preco_referencia', 'preco_referencia', 0,
        output_field=DecimalField(),
    )
    equipamentos = equipamentos.annotate(valor_considerado=valor)
    total_equipamentos = equipamentos.count()
    sem_preco = equipamentos.filter(
        custo_aquisicao=None,
        produto__preco_referencia=None,
        preco_referencia=None,
    ).count()
    precificados = total_equipamentos - sem_preco
    total = equipamentos.aggregate(v=Sum(valor))['v'] or Decimal('0')
    cobertura = round((precificados / total_equipamentos * 100), 1) if total_equipamentos else 0
    valor_medio = total / total_equipamentos if total_equipamentos else Decimal('0')
    categorias = {
        'operacional': equipamentos.filter(finalidade='OPERACIONAL', status__in=['ATIVO', 'EM_USO']).aggregate(v=Sum(valor))['v'] or 0,
        'administrativo': equipamentos.filter(finalidade='ADMINISTRATIVO').aggregate(v=Sum(valor))['v'] or 0,
        'indisponivel': equipamentos.filter(status__in=['SICK', 'MANUTENCAO']).aggregate(v=Sum(valor))['v'] or 0,
        'transito': equipamentos.filter(status='EM_TRANSITO').aggregate(v=Sum(valor))['v'] or 0,
        'baixados': equipamentos.filter(status__in=['BAIXA', 'INATIVO']).aggregate(v=Sum(valor))['v'] or 0,
    }
    equipamentos_ordenados = equipamentos.order_by(
        'regional__nome', 'produto__descricao', 'patrimonio'
    )
    page_obj = Paginator(equipamentos_ordenados, 20).get_page(request.GET.get('page'))
    pode_definir_preco = ComprasAccessPolicy.pode_definir_preco_produto(request.user)
    pode_alterar_preco = ComprasAccessPolicy.pode_alterar_preco_produto(request.user)
    for equipamento in page_obj.object_list:
        equipamento.pode_editar_preco = bool(
            equipamento.produto_id
            and (
                (equipamento.produto.preco_referencia is None and pode_definir_preco)
                or (equipamento.produto.preco_referencia is not None and pode_alterar_preco)
            )
        )
    query_params = request.GET.copy()
    query_params.pop('page', None)
    produtos = Produto.objects.filter(ativo=True)
    if categoria:
        produtos = produtos.filter(categoria=categoria)
    return render(request, 'compras/valores_equipamentos.html', {
        'equipamentos': page_obj.object_list,
        'page_obj': page_obj,
        'query_string': query_params.urlencode(),
        'bases': bases,
        'total': total,
        'total_equipamentos': total_equipamentos,
        'sem_preco': sem_preco,
        'precificados': precificados,
        'cobertura': cobertura,
        'valor_medio': valor_medio,
        'categorias_valor': categorias,
        'pode_editar': ComprasAccessPolicy.pode_editar_precos(request.user),
        'pode_editar_preco_direto': pode_definir_preco or pode_alterar_preco,
        'pode_importar': ComprasAccessPolicy.pode_importar_precos(request.user),
        'pode_catalogo': ComprasAccessPolicy.pode_gerenciar_catalogo(request.user),
        'categorias': Produto.CATEGORIAS,
        'produtos': produtos.order_by('categoria', 'descricao'),
        'filtros': {
            'base': base_id or '', 'categoria': categoria,
            'equipamento': produto_id, 'q': busca,
        },
        'grafico_regionais': list(
            equipamentos.values('regional__nome').annotate(total=Sum(valor)).order_by('-total')[:10]
        ),
        'grafico_categorias': list(
            equipamentos.values('produto__categoria').annotate(total=Sum(valor)).order_by('-total')
        ),
        'grafico_equipamentos': list(
            equipamentos.values('produto__descricao').annotate(total=Sum(valor)).order_by('-total')[:10]
        ),
        'importacao_form': ImportacaoPrecificacaoForm(),
        'origens_preco': Produto.OrigemPreco.choices,
        'fornecedores_preco': (
            FornecedorInsumo.objects.filter(ativo=True).order_by('nome')
            if pode_definir_preco or pode_alterar_preco
            else FornecedorInsumo.objects.none()
        ),
    })


@login_required
@require_POST
def alterar_preco_produto(request, equipamento_id):
    equipamentos = VisibilidadeEstoqueAuditoriaService.ocultar_equipamentos(
        Equipamento.objects.select_related('produto', 'regional__empresa')
    )
    if not request.user.perfil.is_admin:
        equipamentos = equipamentos.filter(regional__in=ComprasAccessPolicy.bases(request.user))
    equipamento = get_object_or_404(equipamentos, pk=equipamento_id)

    retorno = request.POST.get('retorno', '').strip()
    if not url_has_allowed_host_and_scheme(
        retorno,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        retorno = reverse('compras:valores_equipamentos')

    if not equipamento.produto_id:
        messages.error(request, _('O equipamento não possui produto de catálogo vinculado.'))
        return redirect(retorno)
    pode_editar = (
        ComprasAccessPolicy.pode_definir_preco_produto(request.user)
        if equipamento.produto.preco_referencia is None
        else ComprasAccessPolicy.pode_alterar_preco_produto(request.user)
    )
    if not pode_editar:
        raise PermissionDenied

    try:
        valor_texto = request.POST.get('preco_referencia', '').strip().replace(',', '.')
        valor = Decimal(valor_texto)
        observacao = request.POST.get('observacao_preco', '').strip()
        if not observacao:
            raise ValidationError(_('Informe o motivo da alteração do preço.'))
        fornecedor_id = request.POST.get('preco_fornecedor', '').strip()
        fornecedor = None
        if fornecedor_id:
            fornecedor = get_object_or_404(
                FornecedorInsumo.objects.filter(ativo=True), pk=fornecedor_id
            )
        ProdutoPrecoService.definir(
            produto=equipamento.produto,
            usuario=request.user,
            valor=valor,
            origem=request.POST.get('preco_origem', '').strip(),
            fonte=request.POST.get('preco_fonte', '').strip(),
            fornecedor=fornecedor,
            observacao=observacao,
        )
    except (InvalidOperation, ValueError):
        messages.error(request, _('Informe um preço de referência válido.'))
    except (PermissionDenied, ValidationError) as exc:
        messages.error(request, str(exc))
    else:
        messages.success(
            request,
            _('Preço de %(produto)s atualizado com histórico e rastreabilidade.') % {
                'produto': equipamento.produto.descricao,
            },
        )
    return redirect(retorno)


@login_required
def template_precificacao_equipamentos(request):
    if not ComprasAccessPolicy.pode_importar_precos(request.user):
        raise PermissionDenied
    equipamentos = VisibilidadeEstoqueAuditoriaService.ocultar_equipamentos(
        Equipamento.objects.select_related('produto', 'regional')
    )
    if not request.user.perfil.is_admin:
        equipamentos = equipamentos.filter(regional__in=ComprasAccessPolicy.bases(request.user))
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = 'PRECIFICACAO'
    planilha.append([
        'EQUIPAMENTO_ID', 'REGIONAL', 'CATEGORIA', 'EQUIPAMENTO', 'PATRIMONIO',
        'NUMERO_SERIE', 'CUSTO_AQUISICAO', 'PRECO_REFERENCIA', 'ORIGEM_VALOR', 'MOTIVO',
    ])
    for item in equipamentos.order_by('regional__nome', 'produto__descricao', 'patrimonio'):
        referencia = (
            item.produto.preco_referencia
            if item.produto_id and item.produto.preco_referencia is not None
            else item.preco_referencia
        )
        origem = (
            item.produto.preco_origem
            if item.produto_id and item.produto.preco_referencia is not None
            else item.origem_valor
        )
        planilha.append([
            item.pk, item.regional.nome, item.produto.categoria if item.produto else '',
            item.produto.descricao if item.produto else '', item.patrimonio, item.numero_serie,
            item.custo_aquisicao, referencia, origem, '',
        ])
    resposta = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resposta['Content-Disposition'] = 'attachment; filename="template-precificacao-equipamentos.xlsx"'
    workbook.save(resposta)
    return resposta


@login_required
@require_POST
def importar_precificacao_equipamentos(request):
    if not ComprasAccessPolicy.pode_importar_precos(request.user):
        raise PermissionDenied
    form = ImportacaoPrecificacaoForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, _('Planilha de precificação inválida.'))
        return redirect('compras:valores_equipamentos')
    try:
        planilha = load_workbook(form.cleaned_data['arquivo'], read_only=True, data_only=True).active
        linhas = planilha.iter_rows(values_only=True)
        cabecalho = [str(valor or '').strip().upper() for valor in next(linhas)]
        esperadas = ['EQUIPAMENTO_ID', 'CUSTO_AQUISICAO', 'PRECO_REFERENCIA', 'ORIGEM_VALOR', 'MOTIVO']
        if any(coluna not in cabecalho for coluna in esperadas):
            raise ValidationError(_('O cabeçalho da planilha não corresponde ao template oficial.'))
        indices = {nome: cabecalho.index(nome) for nome in esperadas}
        equipamentos_permitidos = VisibilidadeEstoqueAuditoriaService.ocultar_equipamentos(
            Equipamento.objects.all()
        )
        if not request.user.perfil.is_admin:
            equipamentos_permitidos = equipamentos_permitidos.filter(
                regional__in=ComprasAccessPolicy.bases(request.user)
            )
        itens_importacao = []
        linhas_por_id = {}
        for numero, linha in enumerate(linhas, start=2):
            equipamento_id = linha[indices['EQUIPAMENTO_ID']]
            if not equipamento_id:
                continue
            try:
                if isinstance(equipamento_id, bool):
                    raise ValueError
                equipamento_id_convertido = int(equipamento_id)
                if isinstance(equipamento_id, float) and not equipamento_id.is_integer():
                    raise ValueError
            except (TypeError, ValueError):
                raise ValidationError(
                    _('Linha %(numero)s: EQUIPAMENTO_ID deve ser um número inteiro.') % {
                        'numero': numero,
                    }
                )
            if equipamento_id_convertido in linhas_por_id:
                raise ValidationError(_(
                    'Linha %(numero)s: EQUIPAMENTO_ID duplicado; primeira ocorrência na linha '
                    '%(linha_anterior)s.'
                ) % {
                    'numero': numero,
                    'linha_anterior': linhas_por_id[equipamento_id_convertido],
                })
            linhas_por_id[equipamento_id_convertido] = numero
            itens_importacao.append({
                'linha': numero,
                'equipamento_id': equipamento_id_convertido,
                'custo': linha[indices['CUSTO_AQUISICAO']],
                'referencia': linha[indices['PRECO_REFERENCIA']],
                'origem': linha[indices['ORIGEM_VALOR']],
                'motivo': linha[indices['MOTIVO']],
            })

        ids_solicitados = set(linhas_por_id)
        ids_permitidos = set(
            equipamentos_permitidos.filter(pk__in=ids_solicitados).values_list('pk', flat=True)
        )
        ids_invalidos = ids_solicitados - ids_permitidos
        if ids_invalidos:
            primeiro_id = min(ids_invalidos, key=lambda pk: linhas_por_id[pk])
            raise ValidationError(_(
                'Linha %(numero)s: equipamento inexistente ou fora do seu escopo autorizado.'
            ) % {'numero': linhas_por_id[primeiro_id]})

        atualizados, ignorados = AquisicaoService.atualizar_valores_equipamentos_em_lote(
            itens=itens_importacao,
            usuario=request.user,
        )
    except (
        ValidationError, ValueError, TypeError, KeyError, StopIteration,
        BadZipFile, InvalidFileException,
    ) as exc:
        messages.error(request, str(exc))
    else:
        if atualizados:
            messages.success(
                request,
                _(
                    '%(quantidade)s equipamento(s) precificado(s) com sucesso; '
                    '%(ignorados)s linha(s) sem alteração foram ignoradas.'
                ) % {'quantidade': atualizados, 'ignorados': ignorados},
            )
        else:
            messages.info(
                request,
                _('Nenhum valor foi alterado; %(ignorados)s linha(s) já estavam atualizadas.') % {
                    'ignorados': ignorados,
                },
            )
    return redirect('compras:valores_equipamentos')


@login_required
def criar_produto_catalogo(request):
    if not ComprasAccessPolicy.pode_gerenciar_catalogo(request.user):
        raise PermissionDenied
    form = ProdutoForm(request.POST or None, user=request.user)
    if request.method == 'POST' and form.is_valid():
        try:
            with transaction.atomic():
                produto = form.save(commit=False)
                produto.criado_por = request.user
                produto.save()
                preco = form.cleaned_data.get('preco_referencia_inicial')
                if preco is not None:
                    ProdutoPrecoService.definir(
                        produto=produto,
                        usuario=request.user,
                        valor=preco,
                        origem=form.cleaned_data.get('preco_origem'),
                        fonte=form.cleaned_data.get('preco_fonte', ''),
                        fornecedor=form.cleaned_data.get('preco_fornecedor'),
                        observacao='PREÇO DEFINIDO NO CADASTRO DO PRODUTO',
                        somente_inicial=True,
                    )
        except (PermissionDenied, ValidationError) as exc:
            form.add_error(None, exc)
        else:
            messages.success(
                request,
                _('Novo item incluído no catálogo e disponibilizado nos filtros.'),
            )
            destino = (
                'compras:valores_equipamentos'
                if ComprasAccessPolicy.pode_visualizar_valores(request.user)
                else 'estoque:cadastrar_equipamento'
            )
            return redirect(destino)
    return render(request, 'compras/produto_form.html', {'form': form})


@login_required
def resolver_codigo(request):
    codigo = request.GET.get('codigo', '').strip()
    if not codigo:
        return JsonResponse({'erro': 'Informe o código.'}, status=400)
    qs = CodigoCatalogo.objects.filter(codigo=codigo, ativo=True).select_related('produto', 'insumo', 'empresa')
    perfil = request.user.perfil
    if not perfil.is_admin:
        empresas_ids = list(ComprasAccessPolicy.empresas(request.user).values_list('id', flat=True))
        if perfil.empresa_id:
            empresas_ids.append(perfil.empresa_id)
        qs = qs.filter(empresa_id__in=set(empresas_ids))
    registro = qs.first()
    if not registro:
        return JsonResponse({'encontrado': False, 'codigo': codigo}, status=404)
    objeto = registro.produto or registro.insumo
    return JsonResponse({
        'encontrado': True, 'codigo': codigo, 'tipo_codigo': registro.tipo,
        'fator_conversao': str(registro.fator_conversao),
        'tipo_item': 'EQUIPAMENTO' if registro.produto_id else 'INSUMO',
        'id': objeto.pk, 'descricao': objeto.descricao,
    })
