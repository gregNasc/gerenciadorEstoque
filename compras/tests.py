import uuid
from decimal import Decimal
from io import BytesIO

from django.contrib.auth.models import Group, Permission, User
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook

from compras.models import (
    Aquisicao,
    HistoricoPrecoProduto,
    HistoricoValorEquipamento,
    ItemAquisicao,
    RemessaCompra,
)
from compras.services import AquisicaoService, ProdutoPrecoService, RemessaCompraService
from estoque.models import Base, Comunicado, Empresa, Equipamento, Perfil, Produto
from estoque.policies.compras import GruposCorporativos
from estoque.services.sick_service import SickService
from insumos.models import CategoriaInsumo, FornecedorInsumo, Insumo, SaldoInsumoBase
from insumos.services.movimentacao_service import MovimentacaoService
from ordens_servico.models import OrdemServico


@override_settings(
    PASSWORD_HASHERS=['django.contrib.auth.hashers.MD5PasswordHasher'],
)
class ComprasRemessasTests(TestCase):
    def setUp(self):
        self.empresa = Empresa.objects.create(nome='Empresa Compras')
        self.matriz = Base.objects.create(nome='Matriz', empresa=self.empresa)
        self.base = Base.objects.create(nome='Base destino', empresa=self.empresa)
        self.admin = self._usuario('admin_compras', Perfil.Role.ADMIN)
        self.outro_admin = self._usuario('admin_acompanhamento', Perfil.Role.ADMIN)
        self.gestor = self._usuario('gestor_destino', Perfil.Role.GESTOR, self.base)
        self.fornecedor = FornecedorInsumo.objects.create(
            nome='Fornecedor Compras', documento='12345678000190'
        )
        categoria = CategoriaInsumo.objects.create(nome='Categoria Compras')
        self.insumo = Insumo.objects.create(
            descricao='Insumo Compras', categoria=categoria, unidade_medida='UN'
        )
        self.produto = Produto.objects.create(
            codigo='PROD-COMPRA', descricao='Equipamento compra', fabricante='Marca',
            modelo='Modelo', categoria='Coletores',
        )

    @staticmethod
    def _usuario(username, role, base=None):
        user = User.objects.create_user(username, password='senha-forte')
        user.perfil.role = role
        user.perfil.empresa = None if role == Perfil.Role.ADMIN else base.empresa
        user.perfil.save()
        if base:
            user.perfil.regionais.add(base)
        return user

    def _aquisicao(self, quantidade=100):
        aquisicao = AquisicaoService.criar(
            empresa=self.empresa, fornecedor=self.fornecedor, usuario=self.admin,
            numero_documento='NF-123',
            itens=[{
                'tipo_item': ItemAquisicao.Tipo.INSUMO,
                'insumo': self.insumo,
                'quantidade': quantidade,
                'valor_unitario': Decimal('7.50'),
            }],
        )
        AquisicaoService.aprovar(aquisicao, self.admin)
        return aquisicao

    def test_fornecedor_direto_recebe_parcial_e_idempotente(self):
        aquisicao = self._aquisicao()
        with self.captureOnCommitCallbacks(execute=True):
            remessa = RemessaCompraService.criar(
                empresa=self.empresa, fluxo=RemessaCompra.Fluxo.FORNECEDOR_DIRETO,
                aquisicao=aquisicao, base_destino=self.base, usuario=self.admin,
                itens=[{
                    'item_aquisicao': aquisicao.itens.get(),
                    'insumo': self.insumo,
                    'quantidade_prevista': 100,
                    'custo_unitario_snapshot': Decimal('7.50'),
                }],
            )
            RemessaCompraService.enviar(remessa, self.admin)
        chave = uuid.uuid4()
        dados = [{
            'item_id': remessa.itens.get().pk,
            'quantidade_recebida': 98,
            'quantidade_faltante': 2,
        }]
        with self.captureOnCommitCallbacks(execute=True):
            primeiro = RemessaCompraService.confirmar(
                remessa=remessa, usuario=self.gestor, idempotency_key=chave,
                linhas=dados, finalizar=True,
            )
            segundo = RemessaCompraService.confirmar(
                remessa=remessa, usuario=self.gestor, idempotency_key=chave,
                linhas=dados, finalizar=True,
            )
        self.assertEqual(primeiro.pk, segundo.pk)
        remessa.refresh_from_db()
        saldo = SaldoInsumoBase.objects.get(base=self.base, insumo=self.insumo)
        self.assertEqual(saldo.saldo, Decimal('98'))
        self.assertEqual(saldo.custo_medio, Decimal('7.50'))
        self.assertEqual(remessa.status, RemessaCompra.Status.RECEBIDA_PARCIAL)
        self.assertTrue(OrdemServico.objects.filter(chamado_referencia=remessa.protocolo).exists())
        comunicado = Comunicado.objects.filter(dados__remessa_id=remessa.pk).latest('pk')
        self.assertTrue(comunicado.usuarios.filter(pk=self.admin.pk).exists())
        self.assertTrue(comunicado.usuarios.filter(pk=self.outro_admin.pk).exists())
        self.assertTrue(comunicado.usuarios.filter(pk=self.gestor.pk).exists())

    def test_entre_bases_reserva_impede_consumo_e_libera_faltante(self):
        MovimentacaoService.entrada(
            base=self.matriz, insumo=self.insumo, quantidade=100,
            valor_unitario=5, usuario=self.admin,
        )
        remessa = RemessaCompraService.criar(
            empresa=self.empresa, fluxo=RemessaCompra.Fluxo.ENTRE_BASES,
            base_origem=self.matriz, base_destino=self.base, usuario=self.admin,
            itens=[{'insumo': self.insumo, 'quantidade_prevista': 40, 'custo_unitario_snapshot': 5}],
        )
        saldo = SaldoInsumoBase.objects.get(base=self.matriz, insumo=self.insumo)
        self.assertEqual(saldo.saldo_reservado, Decimal('40'))
        with self.assertRaises(ValueError):
            MovimentacaoService.saida(
                base=self.matriz, insumo=self.insumo, quantidade=70, usuario=self.admin,
            )
        RemessaCompraService.enviar(remessa, self.admin)
        RemessaCompraService.confirmar(
            remessa=remessa, usuario=self.gestor, idempotency_key=uuid.uuid4(),
            linhas=[{
                'item_id': remessa.itens.get().pk,
                'quantidade_recebida': 30,
                'quantidade_faltante': 10,
            }],
            finalizar=True,
        )
        origem = SaldoInsumoBase.objects.get(base=self.matriz, insumo=self.insumo)
        destino = SaldoInsumoBase.objects.get(base=self.base, insumo=self.insumo)
        self.assertEqual((origem.saldo, origem.saldo_reservado), (Decimal('70'), Decimal('0')))
        self.assertEqual(destino.saldo, Decimal('30'))

    def test_valor_equipamento_mantem_historico(self):
        equipamento = Equipamento.objects.create(
            produto=self.produto, numero_serie='SERIE-COMPRA', patrimonio='PAT-COMPRA',
            regional=self.base, codigo='EQP-COMPRA',
        )
        AquisicaoService.atualizar_valor_equipamento(
            equipamento=equipamento, usuario=self.admin,
            custo=1000, referencia=1200,
            origem=Equipamento.OrigemValor.INFORMADO_COMPRAS,
            motivo='Valor confirmado na nota', fornecedor=self.fornecedor,
        )
        AquisicaoService.atualizar_valor_equipamento(
            equipamento=equipamento, usuario=self.admin,
            custo=950, referencia=1200,
            origem=Equipamento.OrigemValor.DOCUMENTO_COMPRA,
            motivo='Desconto identificado', fornecedor=self.fornecedor,
        )
        self.assertEqual(HistoricoValorEquipamento.objects.filter(equipamento=equipamento).count(), 2)
        equipamento.refresh_from_db()
        self.assertEqual(equipamento.custo_aquisicao, Decimal('950'))

    def test_importacao_precificacao_ignora_cabecalho_e_atualiza_equipamento(self):
        equipamento = Equipamento.objects.create(
            produto=self.produto, numero_serie='SERIE-IMPORTACAO', patrimonio='PAT-IMPORTACAO',
            regional=self.base, codigo='EQP-IMPORTACAO',
        )
        workbook = Workbook()
        planilha = workbook.active
        planilha.title = 'PRECIFICACAO'
        planilha.append([
            'EQUIPAMENTO_ID', 'REGIONAL', 'CATEGORIA', 'EQUIPAMENTO', 'PATRIMONIO',
            'NUMERO_SERIE', 'CUSTO_AQUISICAO', 'PRECO_REFERENCIA', 'ORIGEM_VALOR', 'MOTIVO',
        ])
        planilha.append([
            equipamento.pk, self.base.nome, self.produto.categoria, self.produto.descricao,
            equipamento.patrimonio, equipamento.numero_serie, 800, 900,
            Equipamento.OrigemValor.INFORMADO_COMPRAS, 'IMPORTAÇÃO DO TEMPLATE',
        ])
        conteudo = BytesIO()
        workbook.save(conteudo)
        arquivo = SimpleUploadedFile(
            'template-precificacao-equipamentos.xlsx', conteudo.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        self.client.force_login(self.admin)
        resposta = self.client.post(
            reverse('compras:importar_precificacao_equipamentos'),
            {'arquivo': arquivo},
        )

        self.assertRedirects(resposta, reverse('compras:valores_equipamentos'))
        equipamento.refresh_from_db()
        self.assertEqual(equipamento.custo_aquisicao, Decimal('800'))
        self.assertEqual(equipamento.preco_referencia, Decimal('900'))
        self.assertEqual(
            HistoricoValorEquipamento.objects.filter(equipamento=equipamento).count(),
            1,
        )

    def test_importacao_ignora_linha_sem_alteracao_mesmo_sem_motivo(self):
        equipamento = Equipamento.objects.create(
            produto=self.produto, numero_serie='SERIE-SEM-ALTERACAO',
            patrimonio='PAT-SEM-ALTERACAO', regional=self.base,
            codigo='EQP-SEM-ALTERACAO', preco_referencia=Decimal('900'),
            origem_valor=Equipamento.OrigemValor.INFORMADO_COMPRAS,
        )
        workbook = Workbook()
        planilha = workbook.active
        planilha.append([
            'EQUIPAMENTO_ID', 'CUSTO_AQUISICAO', 'PRECO_REFERENCIA',
            'ORIGEM_VALOR', 'MOTIVO',
        ])
        planilha.append([
            equipamento.pk, None, 900,
            Equipamento.OrigemValor.INFORMADO_COMPRAS, '',
        ])
        conteudo = BytesIO()
        workbook.save(conteudo)
        arquivo = SimpleUploadedFile(
            'precificacao-sem-alteracao.xlsx', conteudo.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        self.client.force_login(self.admin)
        resposta = self.client.post(
            reverse('compras:importar_precificacao_equipamentos'),
            {'arquivo': arquivo},
        )

        self.assertRedirects(resposta, reverse('compras:valores_equipamentos'))
        self.assertFalse(
            HistoricoValorEquipamento.objects.filter(equipamento=equipamento).exists()
        )

    def test_importacao_em_lote_atualiza_e_comunica_uma_unica_acao(self):
        equipamentos = Equipamento.objects.bulk_create([
            Equipamento(
                produto=self.produto,
                numero_serie=f'SERIE-LOTE-{indice:03d}',
                patrimonio=f'PAT-LOTE-{indice:03d}',
                regional=self.base,
                codigo=f'EQP-LOTE-{indice:03d}',
            )
            for indice in range(120)
        ])
        workbook = Workbook()
        planilha = workbook.active
        planilha.append([
            'EQUIPAMENTO_ID', 'CUSTO_AQUISICAO', 'PRECO_REFERENCIA',
            'ORIGEM_VALOR', 'MOTIVO',
        ])
        for equipamento in equipamentos:
            planilha.append([
                equipamento.pk, None, 750,
                Equipamento.OrigemValor.ESTIMATIVA_MERCADO,
                'Importação de preço padrão',
            ])
        conteudo = BytesIO()
        workbook.save(conteudo)
        arquivo = SimpleUploadedFile(
            'precificacao-em-lote.xlsx', conteudo.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        self.client.force_login(self.admin)
        with self.captureOnCommitCallbacks(execute=True):
            resposta = self.client.post(
                reverse('compras:importar_precificacao_equipamentos'),
                {'arquivo': arquivo},
            )

        self.assertRedirects(resposta, reverse('compras:valores_equipamentos'))
        self.assertEqual(
            Equipamento.objects.filter(
                pk__in=[item.pk for item in equipamentos],
                preco_referencia=Decimal('750'),
                origem_valor=Equipamento.OrigemValor.ESTIMATIVA_MERCADO,
            ).count(),
            120,
        )
        self.assertEqual(
            HistoricoValorEquipamento.objects.filter(
                equipamento_id__in=[item.pk for item in equipamentos]
            ).count(),
            120,
        )
        comunicados = Comunicado.objects.filter(
            dados__acao='VALORES_ATUALIZADOS_EM_LOTE'
        )
        self.assertEqual(comunicados.count(), 1)
        self.assertEqual(comunicados.get().dados['quantidade'], 120)

    def test_paginas_financeiras_e_badge_para_admins_e_envolvidos(self):
        self._aquisicao(1)
        self.client.force_login(self.admin)
        self.assertEqual(self.client.get(reverse('compras:aquisicao_lista')).status_code, 200)
        self.assertEqual(self.client.get(reverse('compras:valores_insumos')).status_code, 200)
        self.assertEqual(self.client.get(reverse('compras:valores_equipamentos')).status_code, 200)

    def test_painel_de_valores_pagina_e_pesquisa_o_inventario_detalhado(self):
        Equipamento.objects.bulk_create([
            Equipamento(
                produto=self.produto, numero_serie=f'SERIE-PAINEL-{indice:02d}',
                patrimonio=f'PAT-PAINEL-{indice:02d}', regional=self.base,
                codigo=f'EQP-PAINEL-{indice:02d}', preco_referencia=indice * 100,
                origem_valor=Equipamento.OrigemValor.ESTIMATIVA_MERCADO,
            )
            for indice in range(1, 22)
        ])
        self.client.force_login(self.admin)

        resposta = self.client.get(reverse('compras:valores_equipamentos'))

        self.assertEqual(resposta.status_code, 200)
        self.assertEqual(resposta.context['total_equipamentos'], 21)
        self.assertEqual(len(resposta.context['equipamentos']), 20)
        self.assertEqual(resposta.context['page_obj'].paginator.num_pages, 2)
        self.assertContains(resposta, 'Consultar inventario detalhado')

        resposta_busca = self.client.get(
            reverse('compras:valores_equipamentos'),
            {'q': 'PAT-PAINEL-21'},
        )
        self.assertEqual(resposta_busca.context['total_equipamentos'], 1)
        self.assertContains(resposta_busca, 'PAT-PAINEL-21')
        self.assertNotContains(resposta_busca, 'PAT-PAINEL-20')

    def test_painel_de_insumos_pagina_pesquisa_e_calcula_disponibilidade(self):
        insumos = Insumo.objects.bulk_create([
            Insumo(
                descricao=f'INSUMO PAINEL {indice:02d}', categoria=self.insumo.categoria,
                unidade_medida='UN', valor_medio=indice,
            )
            for indice in range(1, 22)
        ])
        SaldoInsumoBase.objects.bulk_create([
            SaldoInsumoBase(
                base=self.base, insumo=insumo, saldo=10,
                saldo_reservado=2, custo_medio=indice,
            )
            for indice, insumo in enumerate(insumos, start=1)
        ])
        self.client.force_login(self.admin)

        resposta = self.client.get(reverse('compras:valores_insumos'))

        self.assertEqual(resposta.status_code, 200)
        self.assertEqual(resposta.context['skus'], 21)
        self.assertEqual(len(resposta.context['saldos']), 20)
        self.assertEqual(resposta.context['page_obj'].paginator.num_pages, 2)
        self.assertEqual(resposta.context['saldo_total'], Decimal('210'))
        self.assertEqual(resposta.context['saldo_reservado'], Decimal('42'))
        self.assertEqual(resposta.context['saldo_disponivel'], Decimal('168'))
        self.assertContains(resposta, 'Consultar estoque detalhado')

        resposta_busca = self.client.get(
            reverse('compras:valores_insumos'), {'q': 'INSUMO PAINEL 21'},
        )
        self.assertEqual(resposta_busca.context['skus'], 1)
        self.assertContains(resposta_busca, 'INSUMO PAINEL 21')
        self.assertEqual(
            [item.insumo.descricao for item in resposta_busca.context['saldos']],
            ['INSUMO PAINEL 21'],
        )

    def test_admin_atende_solicitacoes_mas_nao_cria(self):
        self.client.force_login(self.admin)
        self.assertEqual(
            self.client.get(reverse('insumos:criar_solicitacao_insumo')).status_code,
            403,
        )
        self.assertEqual(
            self.client.get(reverse('estoque:criar_solicitacao')).status_code,
            403,
        )
        self.assertEqual(
            self.client.get(reverse('insumos:lista_solicitacoes_insumo')).status_code,
            200,
        )
        self.assertEqual(
            self.client.get(reverse('estoque:caixa_solicitacoes')).status_code,
            200,
        )


class ManutencaoMatrizExclusivaTests(TestCase):
    def setUp(self):
        empresa = Empresa.objects.create(nome='Empresa Matriz')
        self.base = Base.objects.create(nome='Base Matriz', empresa=empresa)
        produto = Produto.objects.create(
            codigo='PROD-MATRIZ', descricao='Coletor Matriz', fabricante='Marca',
            modelo='Modelo', categoria='Coletores',
        )
        self.equipamento = Equipamento.objects.create(
            produto=produto, numero_serie='SERIE-MATRIZ', patrimonio='PAT-MATRIZ',
            regional=self.base, codigo='EQP-MATRIZ',
        )
        self.gestor = User.objects.create_user('gestor_matriz')
        self.gestor.perfil.role = Perfil.Role.GESTOR
        self.gestor.perfil.empresa = empresa
        self.gestor.perfil.save()
        self.gestor.perfil.regionais.add(self.base)
        self.admin_comum = User.objects.create_user('admin_sem_manutencao')
        self.admin_comum.perfil.role = Perfil.Role.ADMIN
        self.admin_comum.perfil.save()
        self.rafael = User.objects.create_user('rafael.ribeiro')
        self.rafael.perfil.role = Perfil.Role.OPERADOR
        self.rafael.perfil.save()
        grupo, _ = Group.objects.get_or_create(name=GruposCorporativos.SICK_MANUTENCAO)
        self.rafael.groups.add(grupo)

    def test_admin_comum_nao_executa_e_tecnico_autorizado_executa(self):
        sick = SickService.marcar_como_sick(
            equipamento_id=self.equipamento.pk, usuario=self.gestor,
            categoria='Hardware', motivo='Falha', observacao='Falha',
        )
        SickService.enviar_para_manutencao(
            sick_id=sick.pk, usuario=self.gestor, destino='Matriz',
        )
        with self.assertRaises(PermissionDenied):
            SickService.confirmar_recebimento(sick_id=sick.pk, usuario=self.admin_comum)
        SickService.confirmar_recebimento(sick_id=sick.pk, usuario=self.rafael)
        sick.refresh_from_db()
        self.assertEqual(sick.etapa, 'RECEBIDO')


@override_settings(PASSWORD_HASHERS=['django.contrib.auth.hashers.MD5PasswordHasher'])
class PrecificacaoProdutoTests(TestCase):
    def setUp(self):
        self.empresa = Empresa.objects.create(nome='Empresa preço produto')
        self.base = Base.objects.create(nome='Base preço produto', empresa=self.empresa)
        self.admin = User.objects.create_user('admin_preco_produto', password='senha-forte')
        self.admin.perfil.role = Perfil.Role.ADMIN
        self.admin.perfil.save(update_fields=['role'])
        self.operador = User.objects.create_user('operador_catalogo', password='senha-forte')
        self.operador.perfil.role = Perfil.Role.OPERADOR
        self.operador.perfil.empresa = self.empresa
        self.operador.perfil.save(update_fields=['role', 'empresa'])
        self.operador.perfil.regionais.add(self.base)
        self.operador.user_permissions.add(
            Permission.objects.get(codename='cadastrar_equipamentos', content_type__app_label='estoque')
        )

    @staticmethod
    def dados_produto(codigo):
        return {
            'codigo': codigo,
            'descricao': 'Coletor de dados',
            'nome_resumido': 'Coletor',
            'fabricante': 'Zebra',
            'modelo': 'TC22',
            'sku_fabricante': '',
            'categoria': 'Coletores',
            'subcategoria': '',
            'unidade_medida': 'UN',
            'quantidade_embalagem': '1',
            'especificacoes_tecnicas': '{}',
            'ativo': 'on',
        }

    def test_admin_cadastra_produto_com_preco_e_historico(self):
        self.client.force_login(self.admin)
        dados = self.dados_produto('PROD-PRECO-1')
        dados.update({
            'preco_referencia_inicial': '1850.00',
            'preco_origem': Produto.OrigemPreco.ESTIMATIVA_MERCADO,
            'preco_fonte': 'Pesquisa de mercado',
        })
        resposta = self.client.post(reverse('compras:criar_produto_catalogo'), dados)
        self.assertRedirects(resposta, reverse('compras:valores_equipamentos'))
        produto = Produto.objects.get(codigo='PROD-PRECO-1')
        self.assertEqual(produto.preco_referencia, Decimal('1850.00'))
        self.assertEqual(produto.historico_precos.count(), 1)

    def test_usuario_sem_preco_nao_define_valores_por_post_manual(self):
        self.client.force_login(self.operador)
        dados = self.dados_produto('PROD-SEM-PERM')
        dados['preco_referencia_inicial'] = '9999.00'
        resposta = self.client.post(reverse('compras:criar_produto_catalogo'), dados)
        self.assertRedirects(resposta, reverse('estoque:cadastrar_equipamento'))
        produto = Produto.objects.get(codigo='PROD-SEM-PERM')
        self.assertIsNone(produto.preco_referencia)

    def test_nova_unidade_reutiliza_preco_sem_alterar_produto(self):
        produto = Produto.objects.create(
            codigo='PROD-HERDA', descricao='Notebook', fabricante='Dell',
            modelo='3420', categoria='Notebooks', preco_referencia=Decimal('2500'),
            preco_origem=Produto.OrigemPreco.ESTIMATIVA_MERCADO,
        )
        equipamento = Equipamento.objects.create(
            produto=produto, numero_serie='SER-HERDA', patrimonio='PAT-HERDA',
            regional=self.base, codigo='EQ-HERDA',
        )
        produto.refresh_from_db()
        self.assertEqual(equipamento.preco_referencia, Decimal('2500'))
        self.assertEqual(produto.preco_referencia, Decimal('2500'))
        self.assertFalse(HistoricoPrecoProduto.objects.filter(produto=produto).exists())

    def test_alteracao_explicita_preserva_historico(self):
        produto = Produto.objects.create(
            codigo='PROD-ALTERA', descricao='Router', fabricante='Cisco',
            modelo='R1', categoria='Routers', preco_referencia=Decimal('500'),
            preco_origem=Produto.OrigemPreco.LEGADO,
        )
        ProdutoPrecoService.definir(
            produto=produto, usuario=self.admin, valor='650',
            origem=Produto.OrigemPreco.INFORMADO_COMPRAS,
            fonte='Cotação', observacao='Atualização controlada', comunicar=False,
        )
        historico = HistoricoPrecoProduto.objects.get(produto=produto)
        self.assertEqual(historico.valor_anterior, Decimal('500'))
        self.assertEqual(historico.valor_novo, Decimal('650'))

    def test_painel_expansivel_altera_preco_sem_planilha(self):
        produto = Produto.objects.create(
            codigo='PROD-PAINEL', descricao='Notebook painel', fabricante='Dell',
            modelo='P1', categoria='Notebooks', preco_referencia=Decimal('1200'),
            preco_origem=Produto.OrigemPreco.LEGADO,
        )
        equipamento = Equipamento.objects.create(
            produto=produto, numero_serie='SER-PAINEL', patrimonio='PAT-PAINEL',
            regional=self.base, codigo='EQ-PAINEL',
        )
        self.client.force_login(self.admin)

        pagina = self.client.get(reverse('compras:valores_equipamentos'))
        self.assertContains(pagina, f'id="detalhe-equipamento-{equipamento.pk}"')
        self.assertContains(
            pagina,
            reverse('compras:alterar_preco_produto', args=[equipamento.pk]),
        )

        resposta = self.client.post(
            reverse('compras:alterar_preco_produto', args=[equipamento.pk]),
            {
                'preco_referencia': '1450.50',
                'preco_origem': Produto.OrigemPreco.INFORMADO_COMPRAS,
                'preco_fonte': 'Cotação direta',
                'observacao_preco': 'Atualização pelo painel detalhado',
                'retorno': reverse('compras:valores_equipamentos'),
            },
        )

        self.assertRedirects(resposta, reverse('compras:valores_equipamentos'))
        produto.refresh_from_db()
        self.assertEqual(produto.preco_referencia, Decimal('1450.50'))
        historico = produto.historico_precos.latest('alterado_em')
        self.assertEqual(historico.valor_anterior, Decimal('1200'))
        self.assertEqual(historico.valor_novo, Decimal('1450.50'))
        self.assertEqual(historico.observacao, 'ATUALIZAÇÃO PELO PAINEL DETALHADO')

    def test_operador_com_permissao_delegada_altera_preco_pelo_painel(self):
        self.operador.user_permissions.add(*Permission.objects.filter(
            content_type__app_label='estoque',
            codename__in=['visualizar_preco_produto', 'alterar_preco_produto'],
        ))
        produto = Produto.objects.create(
            codigo='PROD-DELEGADO', descricao='Coletor delegado', fabricante='Zebra',
            modelo='D1', categoria='Coletores', preco_referencia=Decimal('800'),
            preco_origem=Produto.OrigemPreco.LEGADO,
        )
        equipamento = Equipamento.objects.create(
            produto=produto, numero_serie='SER-DELEGADO', patrimonio='PAT-DELEGADO',
            regional=self.base, codigo='EQ-DELEGADO',
        )
        self.client.force_login(self.operador)

        resposta = self.client.post(
            reverse('compras:alterar_preco_produto', args=[equipamento.pk]),
            {
                'preco_referencia': '875.00',
                'preco_origem': Produto.OrigemPreco.INFORMADO_COMPRAS,
                'observacao_preco': 'Revisão autorizada',
                'retorno': reverse('compras:valores_equipamentos'),
            },
        )

        self.assertRedirects(resposta, reverse('compras:valores_equipamentos'))
        produto.refresh_from_db()
        self.assertEqual(produto.preco_referencia, Decimal('875.00'))
